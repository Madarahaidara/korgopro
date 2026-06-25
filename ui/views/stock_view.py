from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QHeaderView, QMessageBox, QInputDialog,
    QLabel, QFrame, QComboBox, QGroupBox, QDoubleSpinBox, QGridLayout,
    QDateEdit, QTextEdit, QTabWidget, QSplitter, QFormLayout,
    QSpinBox, QCheckBox, QDialog, QDialogButtonBox, QStyle, QProgressDialog, QScrollArea,
    QFileDialog
)
from PySide6.QtCore import Qt, Signal, QDate, QTimer
from PySide6.QtGui import QFont, QColor, QIcon, QPixmap, QAction
from PySide6 import QtGui
from datetime import datetime, timedelta
import json
import os
import csv
import pandas as pd
from sqlalchemy import extract
from sqlalchemy.orm import joinedload
import traceback
import openpyxl 
# Import des modèles SQLAlchemy
from core.database import SessionLocal
from core.models.stock_models import Product, Supplier, InventoryMovement, Expense, ExpenseCategory, PurchaseOrder, StockAlert
from core.models.user import User
from core.models.sale_models import Sale, SaleItem, Customer  # Ajout des modèles de vente

# Import des dialogues d'impression
from utils.print_dialogs import PrintOptionsDialog, PrintHistoryDialog

# Import du SettingsManager
from utils.settings_manager import SettingsManager

class StockView(QWidget):
    """Vue complète de gestion des stocks avec toutes les fonctionnalités"""
    
    # Signaux
    product_added = Signal(dict)
    product_updated = Signal(dict)
    product_deleted = Signal(int)
    inventory_updated = Signal(dict)
    expense_added = Signal(dict)
    settings_changed = Signal(dict)  # Nouveau signal pour les changements de paramètres
    
    def __init__(self, user, settings_manager=None):
        super().__init__()
        self.user = user
        self.settings_manager = settings_manager or SettingsManager()  # Instance du SettingsManager
        
        # Se connecter au signal de changement de paramètres
        if self.settings_manager:
            self.settings_manager.settings_changed.connect(self.on_settings_changed)
        
        self.setWindowTitle("Korgo - Gestion Complète des Stocks")
        # Taille suggérée au démarrage, adaptative ensuite
        self.resize(1200, 750)
        
        # Connexion à la base de données
        self.db_session = SessionLocal()
        
        # Données
        self.products = []
        self.filtered_products = []
        
        # Informations d'entreprise
        self.company_info = self.get_company_info()
        self.currency = self.get_setting('currency', 'FCFA')
        self.tax_rate = self.get_setting('tax_rate', 20.0)
        self.invoice_footer = self.get_setting('invoice_footer', '')
        
        # Initialiser les widgets qui seront créés plus tard
        self._init_widget_references()
        
        # Charger les données initiales
        self.load_data()
        
        self.init_ui()
        self.apply_light_theme()
    
    def get_setting(self, key, default=None):
        """Récupère un paramètre spécifique"""
        if self.settings_manager:
            return self.settings_manager.get_setting(key, default)
        return default
    
    def get_company_info(self):
        """Récupère les informations de l'entreprise"""
        if self.settings_manager:
            return self.settings_manager.get_company_info()
        return {
            'name': 'KORGO',
            'address': '',
            'phone': '',
            'email': '',
            'logo': ''
        }
    
    def on_settings_changed(self, settings):
        """Gérer les changements de paramètres"""
        self.company_info = self.get_company_info()
        self.currency = self.get_setting('currency', 'FCFA')
        self.tax_rate = self.get_setting('tax_rate', 20.0)
        self.invoice_footer = self.get_setting('invoice_footer', '')
        
        # Mettre à jour l'interface si nécessaire
        if hasattr(self, 'sales_total_label') and self.sales_total_label:
            # Mettre à jour les labels avec la nouvelle devise
            self.update_currency_labels()
        
        print(f"Paramètres mis à jour: Devise={self.currency}, TVA={self.tax_rate}%")
    
    def update_currency_labels(self):
        """Mettre à jour les labels avec la devise actuelle"""
        # Cette fonction mettra à jour tous les labels qui affichent des montants
        pass
    
    def apply_light_theme(self):
        """Applique le thème light"""
        import os
        
        # Chercher dans différents chemins possibles
        possible_paths = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "themes", "stock_view.qss"),
        ]
        
        for path in possible_paths:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    self.setStyleSheet(f.read())
                    
                    # Ajouter des styles supplémentaires pour les boutons d'impression et suppression
                    additional_styles = """
                        #printButton {
                            background-color: #3b82f6;
                            color: white;
                            font-weight: bold;
                            padding: 8px 15px;
                            border-radius: 4px;
                        }
                        
                        #printButton:hover {
                            background-color: #2563eb;
                        }
                        
                        #printButton:disabled {
                            background-color: #93c5fd;
                            color: #d1d5db;
                        }
                        
                        #printHistoryButton {
                            background-color: #10b981;
                            color: white;
                            font-weight: bold;
                            padding: 8px 15px;
                            border-radius: 4px;
                        }
                        
                        #printHistoryButton:hover {
                            background-color: #059669;
                        }
                        
                        #deleteSaleButton {
                            background-color: #ef4444;
                            color: white;
                            font-weight: bold;
                            padding: 8px 15px;
                            border-radius: 4px;
                        }
                        
                        #deleteSaleButton:hover {
                            background-color: #dc2626;
                        }
                        
                        #deleteSaleButton:disabled {
                            background-color: #fca5a5;
                            color: #fef2f2;
                        }
                    """
                    self.setStyleSheet(self.styleSheet() + additional_styles)
                    return
            except FileNotFoundError:
                continue
        
        print("Thème non trouvé dans les emplacements possibles")
        # Fallback au thème système
        self.setStyleSheet("")
    
    def _init_widget_references(self):
        """Initialiser les références aux widgets pour éviter AttributeError"""
        # Widgets de l'en-tête
        self.stats_products = None
        self.stats_value = None
        self.stats_low_stock = None
        self.stats_margin = None
        
        # Widgets de l'onglet produits
        self.search_input = None
        self.category_filter = None
        self.show_inactive_checkbox = None  # NOUVEAU
        self.products_table = None
        self.add_product_btn = None
        self.edit_product_btn = None
        self.delete_product_btn = None
        self.reactivate_product_btn = None  # NOUVEAU
        self.inventory_btn = None
        self.refresh_btn = None
        
        # Détails produits
        self.detail_name = None
        self.detail_category = None
        self.detail_quantity = None
        self.detail_purchase = None
        self.detail_sale = None
        self.detail_margin = None
        self.detail_supplier = None
        
        # Widgets de l'onglet mouvements
        self.movements_table = None
        self.movement_start_date = None
        self.movement_end_date = None
        self.movement_type_filter = None
        
        # Widgets de l'onglet dépenses
        self.expenses_table = None
        self.expense_total = None
        self.expense_this_month = None
        self.expense_by_category = None
        self.expense_start_date = None
        self.expense_end_date = None
        
        # Widgets de l'onglet ventes
        self.sales_table = None
        self.sales_total_label = None
        self.sales_count_label = None
        self.sales_start_date = None
        self.sales_end_date = None
        self.sales_status_filter = None
        self.sales_payment_filter = None
        self.delete_sale_btn = None
        
        # Widgets de l'onglet calculs
        self.calc_product_combo = None
        self.calc_purchase_price = None
        self.calc_sale_price = None
        self.calc_expenses = None
        self.calc_quantity = None
        self.calc_gross_price = None
        self.calc_detail_price = None
        self.calc_profit = None
        self.calc_profit_percentage = None
        self.calc_total_value = None
        self.calc_total_cost = None
        self.calc_result_text = None
        
        # Widgets de l'onglet rapports
        self.report_total_value = None
        self.report_total_products = None
        self.report_low_stock = None
        self.report_avg_margin = None
        self.report_total_expenses = None
        self.report_net_profit = None
        self.low_margin_table = None
        
        # Widgets d'impression
        self.print_sale_btn = None
        
        # Nouveaux widgets pour les calculs
        self.calc_suggested_gross = None
        self.calc_suggested_detail = None
        self.calc_min_price = None
        self.calc_margin_target = None
        self.calc_break_even = None
        self.calc_roi = None
    
    def load_data(self):
        """Charger les données depuis la base de données"""
        try:
            # Charger les produits avec leurs fournisseurs (tous, actifs et inactifs)
            self.products = self.db_session.query(Product)\
                .options(joinedload(Product.supplier))\
                .order_by(Product.name).all()
            self.filtered_products = self.products.copy()
            
            # Vérifier et créer des catégories de dépenses par défaut si nécessaire
            self.ensure_default_expense_categories()
            
        except Exception as e:
            print(f"Erreur lors du chargement des données: {e}")
            self.products = []
            self.filtered_products = []
        
    def ensure_default_expense_categories(self):
        """Créer des catégories de dépenses par défaut si elles n'existent pas"""
        default_categories = [
            "Transport",
            "Emballage", 
            "Loyer",
            "Salaires",
            "Fournitures",
            "Marketing",
            "Maintenance",
            "Services",
            "Divers"
        ]
        
        for category_name in default_categories:
            existing = self.db_session.query(ExpenseCategory).filter(
                ExpenseCategory.name == category_name
            ).first()
            
            if not existing:
                new_category = ExpenseCategory(
                    name=category_name,
                    description=f"Catégorie {category_name}"
                )
                self.db_session.add(new_category)
        
        try:
            self.db_session.commit()
        except:
            self.db_session.rollback()
    
    def init_ui(self):
        """Initialisation de l'interface utilisateur"""
        main_layout = QVBoxLayout(self)
        
        # Onglets principaux
        self.tab_widget = QTabWidget()
        self.tab_widget.setObjectName("mainTabWidget")
        
        # Onglet 1: Gestion des produits
        products_tab = self.create_products_tab()
        self.tab_widget.addTab(products_tab, self.style().standardIcon(QStyle.SP_DirOpenIcon), "Produits")
        
        # Onglet 2: Mouvements de stock
        inventory_tab = self.create_inventory_tab()
        self.tab_widget.addTab(inventory_tab, self.style().standardIcon(QStyle.SP_BrowserReload), "Mouvements")
        
        # Onglet 3: Dépenses et frais
        expenses_tab = self.create_expenses_tab()
        self.tab_widget.addTab(expenses_tab, self.style().standardIcon(QStyle.SP_FileDialogDetailedView), "Dépenses")
        
        # Onglet 4: Ventes
        sales_tab = self.create_sales_tab()
        self.tab_widget.addTab(sales_tab, self.style().standardIcon(QStyle.SP_FileDialogContentsView), "Ventes")
        
        # Onglet 5: Calculs et bénéfices
        calculations_tab = self.create_calculations_tab()
        self.tab_widget.addTab(calculations_tab, self.style().standardIcon(QStyle.SP_FileDialogDetailedView), "Calculs")
        
        # Onglet 6: Rapports et bénéfices
        reports_tab = self.create_reports_tab()
        self.tab_widget.addTab(reports_tab, self.style().standardIcon(QStyle.SP_FileDialogInfoView), "Rapports")
        
        main_layout.addWidget(self.tab_widget)
        
        # Charger les données initiales
        self.filter_products()
    
    def create_calculations_tab(self):
        """Créer l'onglet des calculs et bénéfices"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Splitter pour organiser les sections
        splitter = QSplitter(Qt.Vertical)
        
        # Section 1: Calcul de marge par produit
        margin_group = QGroupBox("Calcul de marge par produit")
        margin_layout = QGridLayout(margin_group)
        
        # Sélection du produit
        margin_layout.addWidget(QLabel("Produit:"), 0, 0)
        self.calc_product_combo = QComboBox()
        self.calc_product_combo.currentIndexChanged.connect(self.update_calculations_from_product)
        margin_layout.addWidget(self.calc_product_combo, 0, 1, 1, 2)
        
        # Prix d'achat
        margin_layout.addWidget(QLabel("Prix d'achat (HT):"), 1, 0)
        self.calc_purchase_price = QDoubleSpinBox()
        self.calc_purchase_price.setRange(0, 99999999.99)
        self.calc_purchase_price.setDecimals(2)
        self.calc_purchase_price.setPrefix(f"{self.currency} ")
        self.calc_purchase_price.valueChanged.connect(self.calculate_profit_margin)
        margin_layout.addWidget(self.calc_purchase_price, 1, 1)
        
        # Prix de vente
        margin_layout.addWidget(QLabel("Prix de vente (TTC):"), 1, 2)
        self.calc_sale_price = QDoubleSpinBox()
        self.calc_sale_price.setRange(0, 99999999.99)
        self.calc_sale_price.setDecimals(2)
        self.calc_sale_price.setPrefix(f"{self.currency} ")
        self.calc_sale_price.valueChanged.connect(self.calculate_profit_margin)
        margin_layout.addWidget(self.calc_sale_price, 1, 3)
        
        # Frais supplémentaires
        margin_layout.addWidget(QLabel("Frais supplémentaires:"), 2, 0)
        self.calc_expenses = QDoubleSpinBox()
        self.calc_expenses.setRange(0, 99999999.99)
        self.calc_expenses.setDecimals(2)
        self.calc_expenses.setPrefix(f"{self.currency} ")
        self.calc_expenses.setToolTip("Frais de transport, douane, etc.")
        self.calc_expenses.valueChanged.connect(self.calculate_profit_margin)
        margin_layout.addWidget(self.calc_expenses, 2, 1)
        
        # Quantité
        margin_layout.addWidget(QLabel("Quantité:"), 2, 2)
        self.calc_quantity = QSpinBox()
        self.calc_quantity.setRange(1, 999999)
        self.calc_quantity.setValue(1)
        self.calc_quantity.valueChanged.connect(self.calculate_profit_margin)
        margin_layout.addWidget(self.calc_quantity, 2, 3)
        
        # Section résultats de marge
        results_frame = QFrame()
        results_frame.setObjectName("resultsFrame")
        results_layout = QGridLayout(results_frame)
        
        # Résultats
        self.calc_gross_price = QLabel("Prix de vente brut: 0.00")
        self.calc_gross_price.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_gross_price, 0, 0)
        
        self.calc_detail_price = QLabel("Prix de détail conseillé: 0.00")
        self.calc_detail_price.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_detail_price, 0, 1)
        
        self.calc_profit = QLabel("Bénéfice unitaire: 0.00")
        self.calc_profit.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_profit, 1, 0)
        
        self.calc_profit_percentage = QLabel("Marge bénéficiaire: 0%")
        self.calc_profit_percentage.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_profit_percentage, 1, 1)
        
        self.calc_total_value = QLabel("Valeur totale HT: 0.00")
        self.calc_total_value.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_total_value, 2, 0)
        
        self.calc_total_cost = QLabel("Coût total: 0.00")
        self.calc_total_cost.setObjectName("resultLabel")
        results_layout.addWidget(self.calc_total_cost, 2, 1)
        
        margin_layout.addWidget(results_frame, 3, 0, 1, 4)
        
        splitter.addWidget(margin_group)
        
        # Section 2: Calculs avancés
        advanced_group = QGroupBox("Calculs avancés")
        advanced_layout = QGridLayout(advanced_group)
        
        # Prix minimum recommandé
        advanced_layout.addWidget(QLabel("Marge cible (%):"), 0, 0)
        self.calc_margin_target = QDoubleSpinBox()
        self.calc_margin_target.setRange(0, 1000)
        self.calc_margin_target.setDecimals(1)
        self.calc_margin_target.setValue(30)
        self.calc_margin_target.setSuffix("%")
        self.calc_margin_target.valueChanged.connect(self.calculate_suggested_prices)
        advanced_layout.addWidget(self.calc_margin_target, 0, 1)
        
        # Prix de vente suggéré selon marge cible
        self.calc_suggested_gross = QLabel("Prix de vente suggéré (gros): 0.00")
        self.calc_suggested_gross.setObjectName("suggestedLabel")
        advanced_layout.addWidget(self.calc_suggested_gross, 1, 0, 1, 2)
        
        self.calc_suggested_detail = QLabel("Prix de détail suggéré: 0.00")
        self.calc_suggested_detail.setObjectName("suggestedLabel")
        advanced_layout.addWidget(self.calc_suggested_detail, 2, 0, 1, 2)
        
        # Seuil de rentabilité
        advanced_layout.addWidget(QLabel("Coûts fixes mensuels:"), 3, 0)
        self.calc_fixed_costs = QDoubleSpinBox()
        self.calc_fixed_costs.setRange(0, 99999999.99)
        self.calc_fixed_costs.setDecimals(2)
        self.calc_fixed_costs.setPrefix(f"{self.currency} ")
        self.calc_fixed_costs.valueChanged.connect(self.calculate_break_even)
        advanced_layout.addWidget(self.calc_fixed_costs, 3, 1)
        
        self.calc_break_even = QLabel("Seuil de rentabilité: 0 unités/mois")
        self.calc_break_even.setObjectName("resultLabel")
        advanced_layout.addWidget(self.calc_break_even, 4, 0, 1, 2)
        
        self.calc_roi = QLabel("Retour sur investissement (ROI): 0%")
        self.calc_roi.setObjectName("resultLabel")
        advanced_layout.addWidget(self.calc_roi, 5, 0, 1, 2)
        
        splitter.addWidget(advanced_group)
        
        # Section 3: Résumé détaillé
        summary_group = QGroupBox("Résumé détaillé")
        summary_layout = QVBoxLayout(summary_group)
        
        self.calc_result_text = QTextEdit()
        self.calc_result_text.setReadOnly(True)
        self.calc_result_text.setMaximumHeight(150)
        summary_layout.addWidget(self.calc_result_text)
        
        splitter.addWidget(summary_group)
        
        layout.addWidget(splitter)
        
        # Charger les produits dans la combo
        self.load_products_for_calculations()
        
        return tab
    
    def load_products_for_calculations(self):
        """Charger les produits dans la combo des calculs"""
        if self.calc_product_combo:
            self.calc_product_combo.clear()
            self.calc_product_combo.addItem("-- Personnalisé --", None)
            
            # Ne charger que les produits actifs pour les calculs
            active_products = [p for p in self.products if p.active]
            for product in active_products:
                self.calc_product_combo.addItem(f"{product.code or ''} - {product.name}", product.id)
    
    def update_calculations_from_product(self):
        """Mettre à jour les calculs à partir du produit sélectionné"""
        product_id = self.calc_product_combo.currentData()
        
        if product_id:
            product = self.db_session.query(Product).get(product_id)
            if product:
                self.calc_purchase_price.setValue(product.purchase_price)
                self.calc_sale_price.setValue(product.sale_price)
                self.calculate_profit_margin()
    
    def calculate_profit_margin(self):
        """Calculer la marge bénéficiaire"""
        try:
            purchase = self.calc_purchase_price.value()
            sale = self.calc_sale_price.value()
            expenses = self.calc_expenses.value()
            quantity = self.calc_quantity.value()
            
            if purchase > 0:
                # Coût total par unité avec frais
                total_cost_per_unit = purchase + (expenses / quantity)
                
                # Bénéfice unitaire
                profit_per_unit = sale - total_cost_per_unit
                
                # Marge bénéficiaire (%)
                if sale > 0:
                    profit_margin = (profit_per_unit / sale) * 100
                else:
                    profit_margin = 0
                
                # Marge sur coût (%)
                if total_cost_per_unit > 0:
                    margin_on_cost = (profit_per_unit / total_cost_per_unit) * 100
                else:
                    margin_on_cost = 0
                
                # Valeurs totales
                total_value_ht = purchase * quantity
                total_cost_with_expenses = total_cost_per_unit * quantity
                total_profit = profit_per_unit * quantity
                
                # Mettre à jour les labels
                self.calc_gross_price.setText(f"Prix de vente brut: {sale:,.2f} {self.currency}")
                self.calc_profit.setText(f"Bénéfice unitaire: {profit_per_unit:,.2f} {self.currency}")
                self.calc_profit_percentage.setText(f"Marge bénéficiaire: {profit_margin:.1f}%")
                self.calc_total_value.setText(f"Valeur totale HT: {total_value_ht:,.2f} {self.currency}")
                self.calc_total_cost.setText(f"Coût total: {total_cost_with_expenses:,.2f} {self.currency}")
                
                # Couleurs selon la marge
                if profit_margin > 30:
                    self.calc_profit_percentage.setStyleSheet("color: #10b981; font-weight: bold;")
                elif profit_margin < 10:
                    self.calc_profit_percentage.setStyleSheet("color: #ef4444; font-weight: bold;")
                else:
                    self.calc_profit_percentage.setStyleSheet("")
                
                # Mettre à jour le texte de résultat
                result_text = f"""
=== DÉTAIL DU CALCUL ===

Prix d'achat unitaire (HT): {purchase:,.2f} {self.currency}
Frais supplémentaires: {expenses:,.2f} {self.currency}
Coût total par unité: {total_cost_per_unit:,.2f} {self.currency}
Prix de vente unitaire (TTC): {sale:,.2f} {self.currency}

Bénéfice unitaire: {profit_per_unit:,.2f} {self.currency}
Marge bénéficiaire: {profit_margin:.1f}%
Marge sur coût: {margin_on_cost:.1f}%

Pour {quantity} unités:
- Coût total d'achat: {purchase * quantity:,.2f} {self.currency}
- Frais supplémentaires: {expenses:,.2f} {self.currency}
- Coût total: {total_cost_with_expenses:,.2f} {self.currency}
- Chiffre d'affaires: {sale * quantity:,.2f} {self.currency}
- Bénéfice total: {total_profit:,.2f} {self.currency}
                """
                
                self.calc_result_text.setText(result_text)
                
                # Calculer les prix suggérés
                self.calculate_suggested_prices()
                
                # Calculer le seuil de rentabilité
                self.calculate_break_even()
                
            else:
                self.calc_result_text.setText("Veuillez entrer un prix d'achat valide.")
                
        except Exception as e:
            print(f"Erreur dans calculate_profit_margin: {e}")
            self.calc_result_text.setText(f"Erreur de calcul: {str(e)}")
    
    def calculate_suggested_prices(self):
        """Calculer les prix suggérés selon la marge cible"""
        try:
            purchase = self.calc_purchase_price.value()
            expenses = self.calc_expenses.value()
            quantity = self.calc_quantity.value()
            target_margin = self.calc_margin_target.value()
            
            if purchase > 0:
                total_cost_per_unit = purchase + (expenses / quantity)
                
                # Prix suggéré pour atteindre la marge cible (marge sur prix de vente)
                # Formule: Prix = Coût / (1 - Marge%)
                suggested_price = total_cost_per_unit / (1 - (target_margin / 100))
                
                # Prix de détail suggéré (avec marge plus élevée)
                retail_margin = target_margin + 10
                suggested_retail = total_cost_per_unit / (1 - (retail_margin / 100))
                
                self.calc_suggested_gross.setText(f"Prix de vente suggéré (gros, {target_margin:.0f}% marge): {suggested_price:,.2f} {self.currency}")
                self.calc_suggested_detail.setText(f"Prix de détail suggéré ({retail_margin:.0f}% marge): {suggested_retail:,.2f} {self.currency}")
                
        except Exception as e:
            print(f"Erreur dans calculate_suggested_prices: {e}")
    
    def calculate_break_even(self):
        """Calculer le seuil de rentabilité"""
        try:
            fixed_costs = self.calc_fixed_costs.value()
            purchase = self.calc_purchase_price.value()
            sale = self.calc_sale_price.value()
            expenses = self.calc_expenses.value()
            quantity = self.calc_quantity.value()
            
            if sale > 0 and purchase > 0:
                total_cost_per_unit = purchase + (expenses / quantity)
                profit_per_unit = sale - total_cost_per_unit
                
                if profit_per_unit > 0:
                    break_even_units = fixed_costs / profit_per_unit
                    break_even_revenue = break_even_units * sale
                    
                    self.calc_break_even.setText(f"Seuil de rentabilité: {break_even_units:.0f} unités/mois "
                                                f"({break_even_revenue:,.2f} {self.currency} CA)")
                    
                    # Calcul du ROI
                    if fixed_costs > 0:
                        roi = (profit_per_unit * quantity) / fixed_costs * 100
                        self.calc_roi.setText(f"Retour sur investissement (ROI): {roi:.1f}%")
                    else:
                        self.calc_roi.setText("Retour sur investissement (ROI): N/A")
                else:
                    self.calc_break_even.setText("Seuil de rentabilité: Impossible (marge négative)")
                    self.calc_roi.setText("Retour sur investissement (ROI): N/A")
                    
        except Exception as e:
            print(f"Erreur dans calculate_break_even: {e}")
    
    def create_sales_tab(self):
        """Créer l'onglet d'affichage des ventes avec fonctionnalité d'impression et suppression"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Barre d'outils ventes
        toolbar = QFrame()
        toolbar.setObjectName("toolbarFrame")
        toolbar_layout = QHBoxLayout(toolbar)
        
        # Filtres
        filter_group = QGroupBox("Filtres")
        filter_layout = QHBoxLayout(filter_group)
        
        # Filtre par date
        self.sales_start_date = QDateEdit()
        self.sales_start_date.setDate(QDate.currentDate().addDays(-30))
        self.sales_start_date.setCalendarPopup(True)
        self.sales_start_date.setDisplayFormat("dd/MM/yyyy")
        
        self.sales_end_date = QDateEdit()
        self.sales_end_date.setDate(QDate.currentDate())
        self.sales_end_date.setCalendarPopup(True)
        self.sales_end_date.setDisplayFormat("dd/MM/yyyy")
        
        # Filtre par statut
        self.sales_status_filter = QComboBox()
        self.sales_status_filter.addItem("Tous les statuts")
        self.sales_status_filter.addItems(["COMPLETED", "CANCELLED", "REFUNDED"])
        
        # Filtre par méthode de paiement
        self.sales_payment_filter = QComboBox()
        self.sales_payment_filter.addItem("Tous les paiements")
        self.sales_payment_filter.addItems(["CASH", "CARD", "MOBILE_MONEY", "CHECK"])
        
        # Bouton de filtre
        filter_btn = QPushButton("Filtrer")
        filter_btn.clicked.connect(self.load_sales_table)
        
        # Bouton d'effacement
        clear_btn = QPushButton("Effacer")
        clear_btn.clicked.connect(self.clear_sales_filters)
        
        filter_layout.addWidget(QLabel("Du:"))
        filter_layout.addWidget(self.sales_start_date)
        filter_layout.addWidget(QLabel("Au:"))
        filter_layout.addWidget(self.sales_end_date)
        filter_layout.addWidget(QLabel("Statut:"))
        filter_layout.addWidget(self.sales_status_filter)
        filter_layout.addWidget(QLabel("Paiement:"))
        filter_layout.addWidget(self.sales_payment_filter)
        filter_layout.addWidget(filter_btn)
        filter_layout.addWidget(clear_btn)
        filter_layout.addStretch()
        
        # Statistiques rapides
        stats_group = QGroupBox("Statistiques")
        stats_layout = QHBoxLayout(stats_group)
        
        self.sales_count_label = QLabel("Ventes: 0")
        self.sales_count_label.setObjectName("salesStatsLabel")
        
        self.sales_total_label = QLabel(f"Total: 0.00 {self.currency}")
        self.sales_total_label.setObjectName("salesStatsLabel")
        
        stats_layout.addWidget(self.sales_count_label)
        stats_layout.addWidget(self.sales_total_label)
        stats_layout.addStretch()
        
        toolbar_layout.addWidget(filter_group, 70)
        toolbar_layout.addWidget(stats_group, 30)
        
        layout.addWidget(toolbar)
        
        # Barre d'actions (avec impression et suppression)
        actions_toolbar = QFrame()
        actions_layout = QHBoxLayout(actions_toolbar)
        
        # Bouton d'impression
        self.print_sale_btn = QPushButton("🖨️ Imprimer")
        self.print_sale_btn.setObjectName("printButton")
        self.print_sale_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        self.print_sale_btn.clicked.connect(self.print_selected_sale)
        self.print_sale_btn.setEnabled(False)
        
        # Bouton de suppression
        self.delete_sale_btn = QPushButton("🗑️ Supprimer la vente")
        self.delete_sale_btn.setObjectName("deleteSaleButton")
        self.delete_sale_btn.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        self.delete_sale_btn.clicked.connect(self.delete_selected_sale)
        self.delete_sale_btn.setEnabled(False)
        
        # Bouton rafraîchir
        refresh_btn = QPushButton("🔄 Actualiser")
        refresh_btn.setObjectName("refreshButton")
        refresh_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        refresh_btn.clicked.connect(self.load_sales_table)
        
        # Bouton export
        export_btn = QPushButton("📤 Exporter Excel")
        export_btn.setObjectName("exportButton")
        export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        export_btn.clicked.connect(self.export_sales)
        
        actions_layout.addWidget(self.print_sale_btn)
        actions_layout.addWidget(self.delete_sale_btn)
        actions_layout.addWidget(refresh_btn)
        actions_layout.addWidget(export_btn)
        actions_layout.addStretch()
        
        layout.addWidget(actions_toolbar)
        
        # Tableau des ventes
        self.sales_table = QTableWidget()
        self.sales_table.setObjectName("salesTable")
        
        # Configurer la structure du tableau
        headers = [
            "N° Vente", "Date", "Client", "Articles", "Sous-total", 
            "Remise", "Taxe", "Total", "Paiement", "Statut", "Caissier"
        ]
        
        self.sales_table.setColumnCount(len(headers))
        self.sales_table.setHorizontalHeaderLabels(headers)
        
        # Configuration des en-têtes
        header = self.sales_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # N° Vente
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # Client
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Articles
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Sous-total
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Remise
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Taxe
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Total
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # Paiement
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # Statut
        header.setSectionResizeMode(10, QHeaderView.Stretch)  # Caissier
        
        self.sales_table.setAlternatingRowColors(True)
        self.sales_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.sales_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Connecter la sélection pour activer/désactiver les boutons
        self.sales_table.itemSelectionChanged.connect(self.on_sale_selection_changed)
        
        layout.addWidget(self.sales_table, 1)
        
        # Charger les données initiales
        self.load_sales_table()
        
        return tab
    
    def on_sale_selection_changed(self):
        """Gérer le changement de sélection dans le tableau des ventes"""
        selected = self.sales_table.selectedItems()
        if selected:
            if self.print_sale_btn:
                self.print_sale_btn.setEnabled(True)
            if self.delete_sale_btn:
                self.delete_sale_btn.setEnabled(True)
        else:
            if self.print_sale_btn:
                self.print_sale_btn.setEnabled(False)
            if self.delete_sale_btn:
                self.delete_sale_btn.setEnabled(False)
    
    def get_selected_sale_data(self):
        """Récupérer les données de la vente sélectionnée"""
        selected = self.sales_table.selectedItems()
        if not selected:
            return None
        
        row = selected[0].row()
        
        # Récupérer le numéro de vente
        sale_num_item = self.sales_table.item(row, 0)
        if not sale_num_item:
            return None
        
        sale_number = sale_num_item.text()
        
        # Chercher la vente dans la base de données
        try:
            sale = self.db_session.query(Sale).filter(Sale.sale_number == sale_number).first()
            if not sale:
                return None
            
            # Récupérer le nom du client en utilisant first_name et last_name
            customer_name = ""
            if sale.customer:
                customer_name = f"{sale.customer.first_name or ''} {sale.customer.last_name or ''}".strip()
                if sale.customer.company:
                    if customer_name:
                        customer_name += f" ({sale.customer.company})"
                    else:
                        customer_name = sale.customer.company
            
            return {
                'sale_id': sale.id,
                'sale_number': sale.sale_number,
                'date': sale.sale_date.strftime("%d/%m/%Y %H:%M"),
                'customer': customer_name,
                'total': sale.total_amount,
                'items_count': len(sale.items),
                'cashier': sale.cashier.username if sale.cashier else "Inconnu"
            }
        except Exception as e:
            print(f"Erreur lors de la récupération des données de vente: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    
    def delete_selected_sale(self):
        """Supprimer la vente sélectionnée avec restauration du stock"""
        selected = self.sales_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Sélection", "Veuillez sélectionner une vente à supprimer")
            return
        
        row = selected[0].row()
        sale_number_item = self.sales_table.item(row, 0)
        if not sale_number_item:
            return
        
        sale_number = sale_number_item.text()
        
        try:
            # Récupérer la vente avec ses articles et ses paiements
            from core.models.sale_models import Payment
            
            sale = self.db_session.query(Sale)\
                .options(
                    joinedload(Sale.items),
                    joinedload(Sale.payments)
                )\
                .filter(Sale.sale_number == sale_number)\
                .first()
            
            if not sale:
                QMessageBox.warning(self, "Erreur", "Vente non trouvée!")
                return
            
            # Confirmation
            confirm = QMessageBox.question(
                self, "Confirmation",
                f"Voulez-vous vraiment supprimer la vente #{sale_number} ?\n\n"
                f"Cette action est irréversible et restaurera le stock.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if confirm != QMessageBox.Yes:
                return
            
            # Restaurer le stock
            if sale.sale_status not in ["CANCELLED", "REFUNDED"]:
                for item in sale.items:
                    product = self.db_session.query(Product).get(item.product_id)
                    if product and product.active:  # Vérifier si le produit est actif
                        product.quantity += item.quantity
                        
                        movement = InventoryMovement(
                            product_id=product.id,
                            movement_type="IN",
                            quantity=item.quantity,
                            unit_price=item.unit_price,
                            total_value=item.quantity * item.unit_price,
                            reason="Suppression de vente",
                            reference=f"Annulation vente #{sale_number}",
                            notes=f"Stock restauré suite à suppression de la vente #{sale_number}",
                            user_id=self.user.get('id') if isinstance(self.user, dict) else self.user.id
                        )
                        self.db_session.add(movement)
                    elif product and not product.active:
                        # Produit désactivé, on ne restaure pas le stock
                        QMessageBox.warning(
                            self, "Produit désactivé",
                            f"Le produit '{product.name}' est désactivé, le stock n'a pas été restauré."
                        )
            
            # Supprimer les paiements
            self.db_session.query(Payment).filter(Payment.sale_id == sale.id).delete()
            
            # Supprimer les articles de vente
            for item in sale.items:
                self.db_session.delete(item)
            
            # Supprimer la vente
            self.db_session.delete(sale)
            
            self.db_session.commit()
            
            # Rafraîchir l'affichage
            self.load_sales_table()
            self.update_stats()
            
            QMessageBox.information(
                self, "Succès",
                f"La vente #{sale_number} a été supprimée avec succès !"
            )
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la suppression: {str(e)}")
            import traceback
            traceback.print_exc()

    
    def print_selected_sale(self):
        """Imprimer la vente sélectionnée avec les paramètres d'entreprise"""
        sale_data = self.get_selected_sale_data()
        if not sale_data:
            QMessageBox.warning(self, "Sélection", "Veuillez sélectionner une vente à imprimer")
            return
        
        # Ajouter les informations d'entreprise aux données de vente
        sale_data.update({
            'company_info': self.company_info,
            'tax_rate': self.tax_rate,
            'currency': self.currency,
            'invoice_footer': self.invoice_footer,
            'settings_manager': self.settings_manager
        })
        
        dialog = PrintOptionsDialog(sale_data, self)
        dialog.exec()
    
    def export_sales(self):
        """Exporter les ventes au format Excel"""
        try:
            # Demander où sauvegarder
            default_name = f"ventes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter les ventes", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return  # Annulé par l'utilisateur
            
            # Récupérer toutes les ventes avec leurs clients
            sales = self.db_session.query(Sale)\
                .options(joinedload(Sale.customer))\
                .all()
            
            # Préparer les données
            data = []
            for sale in sales:
                customer_name = ""
                if sale.customer:
                    # Utiliser first_name et last_name
                    customer_name = f"{sale.customer.first_name or ''} {sale.customer.last_name or ''}".strip()
                    if not customer_name and sale.customer.company:
                        customer_name = sale.customer.company
                
                cashier_name = sale.cashier.username if sale.cashier else "Inconnu"
                
                data.append({
                    'N° Vente': sale.sale_number,
                    'Date': sale.sale_date.strftime("%Y-%m-%d %H:%M"),
                    'Client': customer_name,
                    'Sous-total': sale.subtotal,
                    'Remise': sale.discount_amount,
                    'TVA': sale.tax_amount,
                    'Total': sale.total_amount,
                    'Paiement': sale.payment_method,
                    'Statut': sale.sale_status,
                    'Caissier': cashier_name
                })
            
            # Créer un DataFrame pandas
            df = pd.DataFrame(data)
            
            # Exporter en Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ventes', index=False)
                
                # Formater automatiquement les colonnes
                worksheet = writer.sheets['Ventes']
                
                # Ajuster automatiquement la largeur des colonnes
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)
                
                # Formater les colonnes numériques
                from openpyxl.styles import numbers
                for cell in worksheet['D'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                for cell in worksheet['E'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                for cell in worksheet['F'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                for cell in worksheet['G'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"{len(sales)} ventes exportées vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def load_sales_table(self):
        """Charger les ventes dans le tableau"""
        try:
            # Construire la requête
            query = self.db_session.query(Sale).options(
                joinedload(Sale.customer),
                joinedload(Sale.cashier)
            )
            
            # Appliquer les filtres de date
            start_date = self.sales_start_date.date().toPython()
            end_date = self.sales_end_date.date().toPython()
            query = query.filter(Sale.sale_date >= start_date)
            query = query.filter(Sale.sale_date <= end_date + timedelta(days=1))
            
            # Filtrer par statut
            status_filter = self.sales_status_filter.currentText()
            if status_filter != "Tous les statuts":
                query = query.filter(Sale.sale_status == status_filter)
            
            # Filtrer par méthode de paiement
            payment_filter = self.sales_payment_filter.currentText()
            if payment_filter != "Tous les paiements":
                query = query.filter(Sale.payment_method == payment_filter)
            
            # Ordonner par date (plus récent d'abord)
            sales = query.order_by(Sale.sale_date.desc()).limit(500).all()
            
            self.sales_table.setRowCount(len(sales))
            
            total_amount = 0
            item_counts = []
            
            for row, sale in enumerate(sales):
                # N° Vente
                sale_num_item = QTableWidgetItem(sale.sale_number)
                sale_num_item.setTextAlignment(Qt.AlignCenter)
                self.sales_table.setItem(row, 0, sale_num_item)
                
                # Date
                date_str = sale.sale_date.strftime("%d/%m/%Y %H:%M") if sale.sale_date else ""
                date_item = QTableWidgetItem(date_str)
                date_item.setTextAlignment(Qt.AlignCenter)
                self.sales_table.setItem(row, 1, date_item)
                
                # Client
                customer_name = ""
                if sale.customer:
                    # Utiliser first_name et last_name au lieu de full_name
                    customer_name = f"{sale.customer.first_name or ''} {sale.customer.last_name or ''}".strip()
                    if not customer_name and sale.customer.company:
                        customer_name = sale.customer.company
                    elif sale.customer.company:
                        customer_name += f" ({sale.customer.company})"
                elif sale.customer_id:
                    customer_name = f"Client #{sale.customer_id}"
                else:
                    customer_name = "Non spécifié"
                
                customer_item = QTableWidgetItem(customer_name)
                self.sales_table.setItem(row, 2, customer_item)
                
                # Nombre d'articles
                item_count = len(sale.items)
                items_text = f"{item_count} article{'s' if item_count > 1 else ''}"
                items_item = QTableWidgetItem(items_text)
                items_item.setTextAlignment(Qt.AlignCenter)
                self.sales_table.setItem(row, 3, items_item)
                
                # Sous-total
                subtotal_item = QTableWidgetItem(f"{sale.subtotal:,.2f} {self.currency}")
                subtotal_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.sales_table.setItem(row, 4, subtotal_item)
                
                # Remise
                discount_item = QTableWidgetItem(f"{sale.discount_amount:,.2f} {self.currency}")
                discount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.sales_table.setItem(row, 5, discount_item)
                
                # Taxe
                tax_item = QTableWidgetItem(f"{sale.tax_amount:,.2f} {self.currency}")
                tax_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.sales_table.setItem(row, 6, tax_item)
                
                # Total
                total_item = QTableWidgetItem(f"{sale.total_amount:,.2f} {self.currency}")
                total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if sale.total_amount > 0:
                    total_item.setForeground(QColor("#10b981"))
                self.sales_table.setItem(row, 7, total_item)
                
                total_amount += sale.total_amount
                
                # Méthode de paiement
                payment_method_text = {
                    "CASH": "💵 Espèces",
                    "CARD": "💳 Carte",
                    "MOBILE_MONEY": "📱 Mobile Money",
                    "CHECK": "📝 Chèque"
                }.get(sale.payment_method, sale.payment_method)
                
                payment_item = QTableWidgetItem(payment_method_text)
                payment_item.setTextAlignment(Qt.AlignCenter)
                self.sales_table.setItem(row, 8, payment_item)
                
                # Statut
                status_item = QTableWidgetItem(sale.sale_status)
                status_item.setTextAlignment(Qt.AlignCenter)
                
                # Couleur selon le statut
                if sale.sale_status == "COMPLETED":
                    status_item.setForeground(QColor("#10b981"))
                    status_item.setText("✅ Terminée")
                elif sale.sale_status == "CANCELLED":
                    status_item.setForeground(QColor("#ef4444"))
                    status_item.setText("❌ Annulée")
                elif sale.sale_status == "REFUNDED":
                    status_item.setForeground(QColor("#f59e0b"))
                    status_item.setText("↩️ Remboursée")
                
                self.sales_table.setItem(row, 9, status_item)
                
                # Caissier
                cashier_name = sale.cashier.username if sale.cashier else "Inconnu"
                cashier_item = QTableWidgetItem(cashier_name)
                self.sales_table.setItem(row, 10, cashier_item)
                
                # Stocker le nombre d'articles pour les statistiques
                item_counts.append(item_count)
            
            # Mettre à jour les statistiques
            if self.sales_count_label:
                self.sales_count_label.setText(f"Ventes: {len(sales)}")
            
            if self.sales_total_label:
                self.sales_total_label.setText(f"Total: {total_amount:,.2f} {self.currency}")
            
        except Exception as e:
            print(f"Erreur lors du chargement des ventes: {e}")
            import traceback
            traceback.print_exc()
            self.sales_table.setRowCount(0)
            
            if self.sales_count_label:
                self.sales_count_label.setText("Ventes: Erreur")
            
            if self.sales_total_label:
                self.sales_total_label.setText("Total: Erreur")
    
    def clear_sales_filters(self):
        """Effacer les filtres des ventes"""
        self.sales_start_date.setDate(QDate.currentDate().addDays(-30))
        self.sales_end_date.setDate(QDate.currentDate())
        self.sales_status_filter.setCurrentIndex(0)
        self.sales_payment_filter.setCurrentIndex(0)
        self.load_sales_table()
    
    def create_products_tab(self):
        """Créer l'onglet de gestion des produits avec désactivation"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Barre d'outils produits
        toolbar = QFrame()
        toolbar.setObjectName("toolbarFrame")
        toolbar_layout = QHBoxLayout(toolbar)
        
        # Recherche
        search_group = QGroupBox("Recherche")
        search_layout = QHBoxLayout(search_group)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nom, catégorie, code, fournisseur...")
        self.search_input.setObjectName("searchInput")
        self.search_input.textChanged.connect(self.filter_products)
        
        self.category_filter = QComboBox()
        self.category_filter.addItem("Toutes catégories")
        
        # Récupérer les catégories depuis la base de données
        categories = self.db_session.query(Product.category).distinct().all()
        for category in categories:
            if category[0]:
                self.category_filter.addItem(category[0])
        
        self.category_filter.currentTextChanged.connect(self.filter_products)
        
        # NOUVEAU: Filtre pour afficher les produits inactifs
        self.show_inactive_checkbox = QCheckBox("Afficher les produits inactifs")
        self.show_inactive_checkbox.stateChanged.connect(self.filter_products)
        
        search_layout.addWidget(QLabel("Recherche:"))
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(QLabel("Catégorie:"))
        search_layout.addWidget(self.category_filter)
        search_layout.addWidget(self.show_inactive_checkbox)
        
        # Boutons d'actions
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout(actions_group)
        
        self.add_product_btn = QPushButton("Nouveau Produit")
        self.add_product_btn.setObjectName("addButton")
        self.add_product_btn.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        self.add_product_btn.clicked.connect(self.show_add_product_dialog)
        
        self.edit_product_btn = QPushButton("Modifier")
        self.edit_product_btn.setObjectName("editButton")
        self.edit_product_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        self.edit_product_btn.clicked.connect(self.edit_selected_product)
        
        self.delete_product_btn = QPushButton("Désactiver")
        self.delete_product_btn.setObjectName("deleteButton")
        self.delete_product_btn.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        self.delete_product_btn.setToolTip("Désactiver le produit (suppression logique)")
        self.delete_product_btn.clicked.connect(self.deactivate_selected_product)
        
        # NOUVEAU: Bouton Réactiver
        self.reactivate_product_btn = QPushButton("🔄 Réactiver")
        self.reactivate_product_btn.setObjectName("reactivateButton")
        self.reactivate_product_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self.reactivate_product_btn.setToolTip("Réactiver un produit désactivé")
        self.reactivate_product_btn.clicked.connect(self.reactivate_selected_product)
        self.reactivate_product_btn.setEnabled(False)
        
        self.inventory_btn = QPushButton("Inventaire")
        self.inventory_btn.setObjectName("inventoryButton")
        self.inventory_btn.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        self.inventory_btn.clicked.connect(self.show_inventory_dialog)
        
        self.refresh_btn = QPushButton("Actualiser")
        self.refresh_btn.setObjectName("refreshButton")
        self.refresh_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self.refresh_btn.clicked.connect(self.refresh_data)
        
        actions_layout.addWidget(self.add_product_btn)
        actions_layout.addWidget(self.edit_product_btn)
        actions_layout.addWidget(self.delete_product_btn)
        actions_layout.addWidget(self.reactivate_product_btn)
        actions_layout.addWidget(self.inventory_btn)
        actions_layout.addWidget(self.refresh_btn)
        
        toolbar_layout.addWidget(search_group, 70)
        toolbar_layout.addWidget(actions_group, 30)
        
        layout.addWidget(toolbar)
        
        # Tableau des produits
        self.products_table = QTableWidget()
        self.products_table.setObjectName("productsTable")
        
        # Configurer le tableau
        self.setup_products_table_structure()
        
        layout.addWidget(self.products_table, 1)
        
        return tab
    
    def setup_products_table_structure(self):
        """Configurer uniquement la structure du tableau des produits"""
        headers = [
            "ID", "Code", "Nom", "Catégorie", "Quantité", "Prix Achat", 
            "Prix Vente", "Marge %", "Valeur Stock", "Fournisseur", "Statut", "Actions"
        ]
        
        self.products_table.setColumnCount(len(headers))
        self.products_table.setHorizontalHeaderLabels(headers)
        
        header = self.products_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # Nom
        header.setSectionResizeMode(9, QHeaderView.Stretch)  # Fournisseur
        header.setSectionResizeMode(11, QHeaderView.ResizeToContents)  # Actions
        
        self.products_table.setAlternatingRowColors(True)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Connecter la sélection
        self.products_table.itemSelectionChanged.connect(self.on_product_selection_changed)
        
        # Charger les produits
        self.load_products_table()
    
    def on_product_selection_changed(self):
        """Gérer le changement de sélection dans le tableau des produits"""
        selected = self.products_table.selectedItems()
        if selected:
            row = selected[0].row()
            status_item = self.products_table.item(row, 10)
            if status_item:
                is_active = "Actif" in status_item.text()
                # Activer le bouton réactiver seulement si le produit est inactif
                if self.reactivate_product_btn:
                    self.reactivate_product_btn.setEnabled(not is_active)
        else:
            if self.reactivate_product_btn:
                self.reactivate_product_btn.setEnabled(False)
    
    def load_products_table(self):
        """Charger les produits dans le tableau"""
        self.products_table.setRowCount(len(self.filtered_products))
        
        for row, product in enumerate(self.filtered_products):
            # ID
            id_item = QTableWidgetItem(str(product.id))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 0, id_item)
            
            # Code
            code_item = QTableWidgetItem(product.code if product.code else "")
            code_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 1, code_item)
            
            # Nom (avec indication si inactif)
            name_item = QTableWidgetItem(product.name)
            if not product.active:
                name_item.setForeground(QColor("#9ca3af"))  # Gris
                name_item.setToolTip("Produit désactivé")
                name_item.setText(f"{product.name} [INACTIF]")
            self.products_table.setItem(row, 2, name_item)
            
            # Catégorie
            category_item = QTableWidgetItem(product.category)
            category_item.setTextAlignment(Qt.AlignCenter)
            if not product.active:
                category_item.setForeground(QColor("#9ca3af"))
            self.products_table.setItem(row, 3, category_item)
            
            # Quantité
            quantity_item = QTableWidgetItem(str(product.quantity))
            quantity_item.setTextAlignment(Qt.AlignCenter)
            
            if product.active:
                if product.is_low_stock:
                    quantity_item.setForeground(QColor("#ef4444"))
                    quantity_item.setToolTip(f"Stock faible! Minimum: {product.min_stock}")
                elif product.is_out_of_stock:
                    quantity_item.setForeground(QColor("#dc2626"))
                    quantity_item.setToolTip("Rupture de stock!")
            else:
                quantity_item.setForeground(QColor("#9ca3af"))
            
            self.products_table.setItem(row, 4, quantity_item)
            
            # Prix d'achat
            purchase_item = QTableWidgetItem(f"{product.purchase_price:,.2f} {self.currency}")
            purchase_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if not product.active:
                purchase_item.setForeground(QColor("#9ca3af"))
            self.products_table.setItem(row, 5, purchase_item)
            
            # Prix de vente
            sale_item = QTableWidgetItem(f"{product.sale_price:,.2f} {self.currency}")
            sale_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if not product.active:
                sale_item.setForeground(QColor("#9ca3af"))
            self.products_table.setItem(row, 6, sale_item)
            
            # Marge
            margin_item = QTableWidgetItem(f"{product.profit_margin:.1f}%")
            margin_item.setTextAlignment(Qt.AlignCenter)
            
            if product.active:
                if product.profit_margin > 30:
                    margin_item.setForeground(QColor("#10b981"))
                elif product.profit_margin < 10:
                    margin_item.setForeground(QColor("#ef4444"))
            else:
                margin_item.setForeground(QColor("#9ca3af"))
            
            self.products_table.setItem(row, 7, margin_item)
            
            # Valeur stock
            value_item = QTableWidgetItem(f"{product.stock_value:,.2f} {self.currency}")
            value_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if not product.active:
                value_item.setForeground(QColor("#9ca3af"))
            self.products_table.setItem(row, 8, value_item)
            
            # Fournisseur
            supplier_name = product.supplier.name if product.supplier else "N/A"
            supplier_item = QTableWidgetItem(supplier_name)
            if not product.active:
                supplier_item.setForeground(QColor("#9ca3af"))
            self.products_table.setItem(row, 9, supplier_item)
            
            # Statut
            status_item = QTableWidgetItem("✓ Actif" if product.active else "✗ Inactif")
            status_item.setTextAlignment(Qt.AlignCenter)
            
            if product.active:
                status_item.setForeground(QColor("#10b981"))
            else:
                status_item.setForeground(QColor("#ef4444"))
            
            self.products_table.setItem(row, 10, status_item)
            
            # Actions
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(5, 5, 5, 5)
            actions_layout.setSpacing(3)
            
            # Bouton Entrée stock (désactivé pour produits inactifs)
            btn_stock_in = QPushButton("+")
            btn_stock_in.setToolTip("Entrée stock")
            btn_stock_in.setFixedSize(28, 28)
            btn_stock_in.setEnabled(product.active)
            btn_stock_in.clicked.connect(lambda _, p=product: self.stock_operation(p, "IN"))
            
            # Bouton Sortie stock (désactivé pour produits inactifs)
            btn_stock_out = QPushButton("-")
            btn_stock_out.setToolTip("Sortie stock")
            btn_stock_out.setFixedSize(28, 28)
            btn_stock_out.setEnabled(product.active)
            btn_stock_out.clicked.connect(lambda _, p=product: self.stock_operation(p, "OUT"))
            
            # Bouton Ajustement (désactivé pour produits inactifs)
            btn_adjust = QPushButton("A")
            btn_adjust.setToolTip("Ajuster stock")
            btn_adjust.setFixedSize(28, 28)
            btn_adjust.setEnabled(product.active)
            btn_adjust.clicked.connect(lambda _, p=product: self.adjust_stock(p))
            
            # Bouton Modifier
            btn_edit = QPushButton("✎")
            btn_edit.setToolTip("Modifier produit")
            btn_edit.setFixedSize(28, 28)
            btn_edit.clicked.connect(lambda _, p=product: self.edit_product(p))
            
            actions_layout.addWidget(btn_stock_in)
            actions_layout.addWidget(btn_stock_out)
            actions_layout.addWidget(btn_adjust)
            actions_layout.addWidget(btn_edit)
            actions_layout.addStretch()
            
            self.products_table.setCellWidget(row, 11, actions_widget)
    
    def filter_products(self):
        """Filtrer les produits (incluant option d'affichage des inactifs)"""
        search_text = self.search_input.text().lower() if self.search_input else ""
        category_filter = self.category_filter.currentText() if self.category_filter else "Toutes catégories"
        show_inactive = self.show_inactive_checkbox.isChecked() if hasattr(self, 'show_inactive_checkbox') else False
        
        self.filtered_products = []
        
        for product in self.products:
            # Filtrer les produits inactifs si nécessaire
            if not show_inactive and not product.active:
                continue
                
            supplier_name = product.supplier.name.lower() if product.supplier else ""
            
            matches_search = (search_text in product.name.lower() or 
                            search_text in product.category.lower() or
                            search_text in supplier_name or
                            (product.code and search_text in product.code.lower()))
            
            matches_category = (category_filter == "Toutes catégories" or 
                              product.category == category_filter)
            
            if matches_search and matches_category:
                self.filtered_products.append(product)
        
        self.load_products_table()
        self.update_stats()
    
    def update_product_details(self):
        """Mettre à jour les détails du produit sélectionné"""
        pass
    
    def show_add_product_dialog(self):
        """Afficher la boîte de dialogue d'ajout de produit"""
        dialog = ProductDialog(self, self.db_session)
        if dialog.exec():
            try:
                # Le dialogue gère déjà l'ajout en base
                # Recharger les données
                self.load_data()
                self.filter_products()
                
                QMessageBox.information(self, "Succès", "Produit ajouté avec succès!")
                
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def edit_selected_product(self):
        """Modifier le produit sélectionné dans le tableau"""
        selected = self.products_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Sélection", "Veuillez sélectionner un produit à modifier")
            return
            
        row = selected[0].row()
        product_id = int(self.products_table.item(row, 0).text())
        product = self.db_session.query(Product).get(product_id)
        
        if product:
            self.edit_product(product)
    
    def edit_product(self, product):
        """Modifier un produit"""
        dialog = ProductDialog(self, self.db_session, product)
        if dialog.exec():
            try:
                # Le dialogue gère déjà la modification
                # Recharger les données
                self.load_data()
                self.filter_products()
                
                QMessageBox.information(self, "Succès", "Produit modifié avec succès!")
                
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def deactivate_selected_product(self):
        """Désactiver le produit sélectionné (suppression logique)"""
        selected = self.products_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Sélection", "Veuillez sélectionner un produit à désactiver")
            return
            
        row = selected[0].row()
        product_id = int(self.products_table.item(row, 0).text())
        
        product = self.db_session.query(Product).get(product_id)
        if not product:
            QMessageBox.warning(self, "Erreur", "Produit non trouvé!")
            return
        
        if not product.active:
            QMessageBox.warning(self, "Information", "Ce produit est déjà désactivé!")
            return
        
        # Vérifier si le produit a des ventes associées
        from core.models.sale_models import SaleItem
        sales_count = self.db_session.query(SaleItem).filter(SaleItem.product_id == product_id).count()
        
        # Construire le message de confirmation
        message = f"Voulez-vous désactiver le produit '{product.name}' ?\n\n"
        
        if product.quantity > 0:
            message += f"⚠️ Attention : {product.quantity} unités encore en stock.\n"
            message += f"   Ces produits seront considérés comme une perte.\n\n"
        
        if sales_count > 0:
            message += f"📊 Ce produit apparaît dans {sales_count} vente(s).\n"
            message += f"   L'historique des ventes sera conservé.\n\n"
        
        message += "Le produit ne sera plus visible dans les listes actives,\n"
        message += "mais toutes les données historiques seront préservées."
        
        reply = QMessageBox.question(
            self, "Confirmation - Désactivation du produit",
            message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # Désactiver le produit (suppression logique)
                product.active = False
                
                # Si le produit a du stock, le passer en perte
                if product.quantity > 0:
                    # Créer un mouvement de perte pour le stock restant
                    loss_movement = InventoryMovement(
                        product_id=product.id,
                        movement_type="LOSS",
                        quantity=product.quantity,
                        unit_price=product.purchase_price,
                        total_value=product.quantity * product.purchase_price,
                        reason="Produit désactivé",
                        notes=f"Stock soldé suite à désactivation du produit. Quantité: {product.quantity} unités",
                        user_id=self.user.get('id') if isinstance(self.user, dict) else self.user.id
                    )
                    self.db_session.add(loss_movement)
                    
                    # Mettre la quantité à 0
                    old_quantity = product.quantity
                    product.quantity = 0
                    
                    QMessageBox.information(
                        self, "Stock soldé", 
                        f"{old_quantity} unité(s) ont été passées en perte.\n"
                        f"Valeur totale: {old_quantity * product.purchase_price:,.2f} {self.currency}"
                    )
                
                self.db_session.commit()
                
                # Recharger les données
                self.load_data()
                self.filter_products()
                self.update_stats()
                
                QMessageBox.information(
                    self, "Succès", 
                    f"Produit '{product.name}' désactivé avec succès !\n\n"
                    f"✓ Le produit n'apparaîtra plus dans les listes actives\n"
                    f"✓ L'historique des ventes est conservé\n"
                    f"✓ Les mouvements de stock sont archivés"
                )
                
            except Exception as e:
                self.db_session.rollback()
                QMessageBox.critical(self, "Erreur", f"Erreur lors de la désactivation: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def reactivate_selected_product(self):
        """Réactiver un produit désactivé"""
        selected = self.products_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Sélection", "Veuillez sélectionner un produit à réactiver")
            return
            
        row = selected[0].row()
        product_id = int(self.products_table.item(row, 0).text())
        
        product = self.db_session.query(Product).get(product_id)
        if not product:
            QMessageBox.warning(self, "Erreur", "Produit non trouvé!")
            return
        
        if product.active:
            QMessageBox.warning(self, "Information", "Ce produit est déjà actif!")
            return
        
        reply = QMessageBox.question(
            self, "Confirmation - Réactivation",
            f"Voulez-vous réactiver le produit '{product.name}' ?\n\n"
            f"Le produit redeviendra visible dans les listes actives.\n"
            f"Note: Le stock actuel est de 0 unité.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                product.active = True
                self.db_session.commit()
                
                # Recharger les données
                self.load_data()
                self.filter_products()
                self.update_stats()
                
                QMessageBox.information(self, "Succès", f"Produit '{product.name}' réactivé avec succès!")
                
            except Exception as e:
                self.db_session.rollback()
                QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def stock_operation(self, product, operation_type):
        """Effectuer une opération de stock (entrée/sortie)"""
        if not product.active:
            QMessageBox.warning(self, "Produit inactif", 
                              "Ce produit est désactivé. Vous ne pouvez pas effectuer d'opérations de stock.")
            return
            
        dialog = StockOperationDialog(self, product, operation_type)
        if dialog.exec():
            quantity = dialog.quantity_spinbox.value()
            notes = dialog.notes_text.toPlainText()
            unit_price = dialog.price_spinbox.value() if hasattr(dialog, 'price_spinbox') else product.purchase_price
            
            try:
                # Créer le mouvement
                movement = InventoryMovement(
                    product_id=product.id,
                    movement_type=operation_type,
                    quantity=quantity,
                    unit_price=unit_price,
                    total_value=quantity * unit_price,
                    reason="Opération manuelle",
                    notes=notes,
                    user_id=self.user.get('id') if isinstance(self.user, dict) else self.user.id
                )
                
                self.db_session.add(movement)
                
                # Mettre à jour la quantité du produit
                if operation_type == "IN":
                    product.quantity += quantity
                elif operation_type == "OUT":
                    if quantity > product.quantity:
                        QMessageBox.warning(self, "Erreur", "Quantité insuffisante en stock!")
                        self.db_session.rollback()
                        return
                    product.quantity -= quantity
                
                self.db_session.commit()
                
                # Recharger les données
                self.load_data()
                self.filter_products()
                self.load_movements_table()
                
                QMessageBox.information(self, "Succès", 
                                      f"Stock {'augmenté' if operation_type == 'IN' else 'diminué'} de {quantity} unités")
                
            except Exception as e:
                self.db_session.rollback()
                QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def adjust_stock(self, product):
        """Ajuster le stock d'un produit"""
        if not product.active:
            QMessageBox.warning(self, "Produit inactif", 
                              "Ce produit est désactivé. Vous ne pouvez pas ajuster le stock.")
            return
            
        dialog = AdjustStockDialog(self, product)
        if dialog.exec():
            new_quantity = dialog.quantity_spinbox.value()
            notes = dialog.notes_text.toPlainText()
            
            difference = new_quantity - product.quantity
            
            if difference == 0:
                QMessageBox.information(self, "Information", "Aucun ajustement nécessaire")
                return
            
            try:
                # Créer le mouvement d'ajustement
                movement = InventoryMovement(
                    product_id=product.id,
                    movement_type="ADJUST",
                    quantity=abs(difference),
                    unit_price=product.purchase_price,
                    total_value=abs(difference) * product.purchase_price,
                    reason="Ajustement manuel",
                    notes=f"{notes}\nDifférence: {'+' if difference > 0 else ''}{difference}",
                    user_id=self.user.get('id') if isinstance(self.user, dict) else self.user.id
                )
                
                self.db_session.add(movement)
                
                # Mettre à jour la quantité
                product.quantity = new_quantity
                self.db_session.commit()
                
                # Recharger les données
                self.load_data()
                self.filter_products()
                self.load_movements_table()
                
                QMessageBox.information(self, "Succès", 
                                      f"Stock ajusté à {new_quantity} unités (différence: {'+' if difference > 0 else ''}{difference})")
                
            except Exception as e:
                self.db_session.rollback()
                QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def show_inventory_dialog(self):
        """Afficher la boîte de dialogue d'inventaire"""
        dialog = InventoryDialog(self, self.db_session, self.user)
        if dialog.exec():
            # Recharger les données
            self.load_data()
            self.filter_products()
            self.load_movements_table()
            self.update_stats()
            
            QMessageBox.information(self, "Succès", "Inventaire mis à jour avec succès!")
    
    def create_inventory_tab(self):
        """Créer l'onglet des mouvements de stock"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Barre d'outils inventaire
        toolbar = QFrame()
        toolbar_layout = QHBoxLayout(toolbar)
        
        # Filtres de date
        date_group = QGroupBox("Filtres")
        date_layout = QHBoxLayout(date_group)
        
        self.movement_start_date = QDateEdit()
        self.movement_start_date.setDate(QDate.currentDate().addDays(-30))
        self.movement_start_date.setCalendarPopup(True)
        
        self.movement_end_date = QDateEdit()
        self.movement_end_date.setDate(QDate.currentDate())
        self.movement_end_date.setCalendarPopup(True)
        
        self.movement_type_filter = QComboBox()
        self.movement_type_filter.addItem("Tous les types")
        self.movement_type_filter.addItems(["IN", "OUT", "ADJUST", "LOSS", "RETURN"])
        
        filter_btn = QPushButton("Filtrer")
        filter_btn.clicked.connect(self.load_movements_table)
        
        clear_btn = QPushButton("Effacer")
        clear_btn.clicked.connect(self.clear_movement_filters)
        
        date_layout.addWidget(QLabel("Du:"))
        date_layout.addWidget(self.movement_start_date)
        date_layout.addWidget(QLabel("Au:"))
        date_layout.addWidget(self.movement_end_date)
        date_layout.addWidget(QLabel("Type:"))
        date_layout.addWidget(self.movement_type_filter)
        date_layout.addWidget(filter_btn)
        date_layout.addWidget(clear_btn)
        date_layout.addStretch()
        
        # Boutons d'action
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout(actions_group)
        
        btn_new_movement = QPushButton("Nouveau mouvement")
        btn_new_movement.setObjectName("addButton")
        btn_new_movement.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        btn_new_movement.clicked.connect(self.show_new_movement_dialog)
        
        btn_adjust_inventory = QPushButton("Ajuster inventaire")
        btn_adjust_inventory.setObjectName("inventoryButton")
        btn_adjust_inventory.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        btn_adjust_inventory.clicked.connect(self.show_inventory_dialog)
        
        btn_export_movements = QPushButton("Exporter Excel")
        btn_export_movements.setObjectName("exportButton")
        btn_export_movements.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        btn_export_movements.clicked.connect(self.export_movements)
        
        actions_layout.addWidget(btn_new_movement)
        actions_layout.addWidget(btn_adjust_inventory)
        actions_layout.addWidget(btn_export_movements)
        actions_layout.addStretch()
        
        toolbar_layout.addWidget(date_group, 70)
        toolbar_layout.addWidget(actions_group, 30)
        
        layout.addWidget(toolbar)
        
        # Tableau des mouvements
        self.movements_table = QTableWidget()
        self.movements_table.setObjectName("movementsTable")
        self.setup_movements_table()
        
        layout.addWidget(self.movements_table, 1)
        
        return tab
    
    def setup_movements_table(self):
        """Configurer le tableau des mouvements"""
        headers = ["ID", "Date", "Produit", "Type", "Quantité", "Prix unitaire", 
                  "Valeur totale", "Raison", "Utilisateur", "Notes"]
        
        self.movements_table.setColumnCount(len(headers))
        self.movements_table.setHorizontalHeaderLabels(headers)
        
        header = self.movements_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # Produit
        header.setSectionResizeMode(7, QHeaderView.Stretch)  # Raison
        header.setSectionResizeMode(9, QHeaderView.Stretch)  # Notes
        
        # Charger les données
        self.load_movements_table()
    
    def load_movements_table(self):
        """Charger les mouvements dans le tableau"""
        try:
            # Construire la requête avec filtres
            query = self.db_session.query(InventoryMovement)\
            .options(
                joinedload(InventoryMovement.product),
                joinedload(InventoryMovement.user)
            )
            
            # Filtre par date
            if hasattr(self, 'movement_start_date') and self.movement_start_date:
                start_date = self.movement_start_date.date().toPython()
                end_date = self.movement_end_date.date().toPython()
                query = query.filter(InventoryMovement.date >= start_date)
                query = query.filter(InventoryMovement.date <= end_date + timedelta(days=1))
            
            # Filtre par type
            if hasattr(self, 'movement_type_filter') and self.movement_type_filter:
                movement_type = self.movement_type_filter.currentText()
                if movement_type != "Tous les types":
                    query = query.filter(InventoryMovement.movement_type == movement_type)
            
            # Ordonner par date (plus récent d'abord)
            movements = query.order_by(InventoryMovement.date.desc()).limit(1000).all()
            
            self.movements_table.setRowCount(len(movements))
            
            for row, movement in enumerate(movements):
                # ID
                id_item = QTableWidgetItem(str(movement.id))
                id_item.setTextAlignment(Qt.AlignCenter)
                self.movements_table.setItem(row, 0, id_item)
                
                # Date
                date_item = QTableWidgetItem(movement.date.strftime("%d/%m/%Y %H:%M"))
                date_item.setTextAlignment(Qt.AlignCenter)
                self.movements_table.setItem(row, 1, date_item)
                
                # Produit
                product_name = f"{movement.product.code} - {movement.product.name}" if movement.product and movement.product.code else (movement.product.name if movement.product else "N/A")
                if movement.product and not movement.product.active:
                    product_name += " [INACTIF]"
                product_item = QTableWidgetItem(product_name)
                self.movements_table.setItem(row, 2, product_item)
                
                # Type
                type_item = QTableWidgetItem(movement.movement_type)
                type_item.setTextAlignment(Qt.AlignCenter)
                
                # Couleur selon le type
                if movement.movement_type == "IN":
                    type_item.setForeground(QColor("#10b981"))
                elif movement.movement_type == "OUT":
                    type_item.setForeground(QColor("#ef4444"))
                elif movement.movement_type == "ADJUST":
                    type_item.setForeground(QColor("#f59e0b"))
                elif movement.movement_type == "LOSS":
                    type_item.setForeground(QColor("#ef4444"))
                
                self.movements_table.setItem(row, 3, type_item)
                
                # Quantité
                qty_item = QTableWidgetItem(f"{movement.quantity:,}")
                qty_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.movements_table.setItem(row, 4, qty_item)
                
                # Prix unitaire
                price_item = QTableWidgetItem(f"{movement.unit_price:,.2f} {self.currency}" if movement.unit_price else "N/A")
                price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.movements_table.setItem(row, 5, price_item)
                
                # Valeur totale
                value_item = QTableWidgetItem(f"{movement.total_value:,.2f} {self.currency}" if movement.total_value else "N/A")
                value_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.movements_table.setItem(row, 6, value_item)
                
                # Raison
                reason_item = QTableWidgetItem(movement.reason or "")
                self.movements_table.setItem(row, 7, reason_item)
                
                # Utilisateur
                username = movement.user.username if movement.user else "Système"
                user_item = QTableWidgetItem(username)
                self.movements_table.setItem(row, 8, user_item)
                
                # Notes
                notes_item = QTableWidgetItem(movement.notes or "")
                self.movements_table.setItem(row, 9, notes_item)
                
        except Exception as e:
            print(f"Erreur lors du chargement des mouvements: {e}")
            self.movements_table.setRowCount(0)
    
    def clear_movement_filters(self):
        """Effacer les filtres de mouvements"""
        if hasattr(self, 'movement_start_date') and self.movement_start_date:
            self.movement_start_date.setDate(QDate.currentDate().addDays(-30))
            self.movement_end_date.setDate(QDate.currentDate())
            self.movement_type_filter.setCurrentIndex(0)
            self.load_movements_table()
    
    def show_new_movement_dialog(self):
        """Afficher la boîte de dialogue de nouveau mouvement"""
        dialog = NewMovementDialog(self, self.db_session, self.user)
        if dialog.exec():
            # Recharger les données
            self.load_data()
            self.filter_products()
            self.load_movements_table()
            self.update_stats()
            
            QMessageBox.information(self, "Succès", "Mouvement enregistré avec succès!")
    
    def export_movements(self):
        """Exporter les mouvements en Excel"""
        try:
            default_name = f"mouvements_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter les mouvements", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            # Récupérer tous les mouvements
            movements = self.db_session.query(InventoryMovement).all()
            
            # Préparer les données
            data = []
            for movement in movements:
                product_name = movement.product.name if movement.product else f"ID:{movement.product_id}"
                username = movement.user.username if movement.user else "Système"
                
                data.append({
                    'ID': movement.id,
                    'Date': movement.date.strftime("%Y-%m-%d %H:%M:%S"),
                    'Produit ID': movement.product_id,
                    'Produit': product_name,
                    'Type': movement.movement_type,
                    'Quantité': movement.quantity,
                    'Prix unitaire': movement.unit_price or 0,
                    'Valeur totale': movement.total_value or 0,
                    'Raison': movement.reason or "",
                    'Utilisateur': username,
                    'Notes': movement.notes or ""
                })
            
            # Exporter en Excel
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Mouvements', index=False)
                
                # Formater automatiquement les colonnes
                worksheet = writer.sheets['Mouvements']
                
                # Ajuster automatiquement la largeur des colonnes
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)
                
                # Formater les colonnes numériques
                from openpyxl.styles import numbers
                for cell in worksheet['G'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                for cell in worksheet['H'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"{len(movements)} mouvements exportés vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur: {str(e)}")
    
    def create_expenses_tab(self):
        """Créer l'onglet des dépenses"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Barre d'outils dépenses
        toolbar = QFrame()
        toolbar_layout = QHBoxLayout(toolbar)
        
        # Filtres de date
        date_group = QGroupBox("Filtres")
        date_layout = QHBoxLayout(date_group)
        
        self.expense_start_date = QDateEdit()
        self.expense_start_date.setDate(QDate.currentDate().addDays(-30))
        self.expense_start_date.setCalendarPopup(True)
        
        self.expense_end_date = QDateEdit()
        self.expense_end_date.setDate(QDate.currentDate())
        self.expense_end_date.setCalendarPopup(True)
        
        filter_btn = QPushButton("Filtrer")
        filter_btn.clicked.connect(self.load_expenses_table)
        
        clear_btn = QPushButton("Effacer")
        clear_btn.clicked.connect(self.clear_expense_filters)
        
        date_layout.addWidget(QLabel("Du:"))
        date_layout.addWidget(self.expense_start_date)
        date_layout.addWidget(QLabel("Au:"))
        date_layout.addWidget(self.expense_end_date)
        date_layout.addWidget(filter_btn)
        date_layout.addWidget(clear_btn)
        date_layout.addStretch()
        
        # Boutons d'action
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout(actions_group)
        
        btn_new_expense = QPushButton("Nouvelle dépense")
        btn_new_expense.setObjectName("addButton")
        btn_new_expense.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        btn_new_expense.clicked.connect(self.show_new_expense_dialog)
        
        btn_expense_report = QPushButton("Rapport")
        btn_expense_report.setObjectName("reportButton")
        btn_expense_report.setIcon(self.style().standardIcon(QStyle.SP_FileDialogInfoView))
        btn_expense_report.clicked.connect(self.show_expense_report)
        
        btn_export_expenses = QPushButton("Exporter Excel")
        btn_export_expenses.setObjectName("exportButton")
        btn_export_expenses.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        btn_export_expenses.clicked.connect(self.export_expenses)
        
        actions_layout.addWidget(btn_new_expense)
        actions_layout.addWidget(btn_expense_report)
        actions_layout.addWidget(btn_export_expenses)
        actions_layout.addStretch()
        
        toolbar_layout.addWidget(date_group, 70)
        toolbar_layout.addWidget(actions_group, 30)
        
        layout.addWidget(toolbar)
        
        # Tableau des dépenses
        self.expenses_table = QTableWidget()
        self.expenses_table.setObjectName("expensesTable")
        
        # Configurer la structure
        headers = ["ID", "Date", "Catégorie", "Montant", "Description", "Moyen de paiement", "Référence", "Utilisateur"]
        self.expenses_table.setColumnCount(len(headers))
        self.expenses_table.setHorizontalHeaderLabels(headers)
        
        header = self.expenses_table.horizontalHeader()
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # Description
        
        layout.addWidget(self.expenses_table, 1)
        
        # Résumé des dépenses
        summary_frame = QFrame()
        summary_frame.setObjectName("summaryFrame")
        summary_layout = QHBoxLayout(summary_frame)
        
        self.expense_total = QLabel(f"Total: 0.00 {self.currency}")
        self.expense_total.setObjectName("summaryLabel")
        
        self.expense_this_month = QLabel(f"Ce mois: 0.00 {self.currency}")
        self.expense_this_month.setObjectName("summaryLabel")
        
        self.expense_by_category = QLabel("Catégorie principale: -")
        self.expense_by_category.setObjectName("summaryLabel")
        
        summary_layout.addWidget(self.expense_total)
        summary_layout.addWidget(self.expense_this_month)
        summary_layout.addWidget(self.expense_by_category)
        summary_layout.addStretch()
        
        layout.addWidget(summary_frame)
        
        # Charger les données
        self.load_expenses_table()
        
        return tab
    
    def load_expenses_table(self):
        """Charger les dépenses dans le tableau"""
        try:
            # Construire la requête
            query = self.db_session.query(Expense).join(ExpenseCategory)
            
            # Filtre par date
            if hasattr(self, 'expense_start_date') and self.expense_start_date:
                start_date = self.expense_start_date.date().toPython()
                end_date = self.expense_end_date.date().toPython()
                query = query.filter(Expense.date >= start_date)
                query = query.filter(Expense.date <= end_date + timedelta(days=1))
            
            # Ordonner par date (plus récent d'abord)
            expenses = query.order_by(Expense.date.desc()).all()
            
            self.expenses_table.setRowCount(len(expenses))
            
            total = 0
            categories = {}
            
            for row, expense in enumerate(expenses):
                # ID
                id_item = QTableWidgetItem(str(expense.id))
                id_item.setTextAlignment(Qt.AlignCenter)
                self.expenses_table.setItem(row, 0, id_item)
                
                # Date
                date_item = QTableWidgetItem(expense.date.strftime("%d/%m/%Y"))
                date_item.setTextAlignment(Qt.AlignCenter)
                self.expenses_table.setItem(row, 1, date_item)
                
                # Catégorie
                category_item = QTableWidgetItem(expense.category.name)
                self.expenses_table.setItem(row, 2, category_item)
                
                # Montant
                amount_item = QTableWidgetItem(f"{expense.amount:,.2f} {self.currency}")
                amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.expenses_table.setItem(row, 3, amount_item)
                
                total += expense.amount
                
                # Compter par catégorie
                if expense.category.name not in categories:
                    categories[expense.category.name] = 0
                categories[expense.category.name] += expense.amount
                
                # Description
                desc_item = QTableWidgetItem(expense.description)
                self.expenses_table.setItem(row, 4, desc_item)
                
                # Moyen de paiement
                payment_item = QTableWidgetItem(expense.payment_method or "Non spécifié")
                self.expenses_table.setItem(row, 5, payment_item)
                
                # Référence
                ref_item = QTableWidgetItem(expense.reference or "")
                self.expenses_table.setItem(row, 6, ref_item)
                
                # Utilisateur
                username = expense.user.username if expense.user else "Système"
                user_item = QTableWidgetItem(username)
                self.expenses_table.setItem(row, 7, user_item)
            
            # Mettre à jour les résumés
            if self.expense_total:
                self.expense_total.setText(f"Total: {total:,.2f} {self.currency}")
            
            # Calculer les dépenses du mois en cours
            if self.expense_this_month:
                current_month = datetime.now().month
                month_expenses = self.db_session.query(Expense).filter(
                    extract('month', Expense.date) == current_month
                ).all()
                month_total = sum(e.amount for e in month_expenses)
                self.expense_this_month.setText(f"Ce mois: {month_total:,.2f} {self.currency}")
            
            # Catégorie principale
            if self.expense_by_category and categories:
                main_category = max(categories.items(), key=lambda x: x[1])
                self.expense_by_category.setText(f"Catégorie principale: {main_category[0]} ({main_category[1]:,.2f} {self.currency})")
            
        except Exception as e:
            print(f"Erreur lors du chargement des dépenses: {e}")
            self.expenses_table.setRowCount(0)
    
    def clear_expense_filters(self):
        """Effacer les filtres de dépenses"""
        if hasattr(self, 'expense_start_date') and self.expense_start_date:
            self.expense_start_date.setDate(QDate.currentDate().addDays(-30))
            self.expense_end_date.setDate(QDate.currentDate())
            self.load_expenses_table()
    
    def show_new_expense_dialog(self):
        """Afficher la boîte de dialogue de nouvelle dépense"""
        dialog = ExpenseDialog(self, self.db_session, self.user)
        if dialog.exec():
            # Recharger les données
            self.load_expenses_table()
            self.update_stats()
            
            QMessageBox.information(self, "Succès", "Dépense enregistrée avec succès!")
    
    def show_expense_report(self):
        """Afficher le rapport des dépenses"""
        try:
            # Calculer les totaux par catégorie
            categories = {}
            all_expenses = self.db_session.query(Expense).all()
            
            for expense in all_expenses:
                category_name = expense.category.name
                if category_name not in categories:
                    categories[category_name] = {"count": 0, "total": 0}
                
                categories[category_name]["count"] += 1
                categories[category_name]["total"] += expense.amount
            
            # Créer un dialogue personnalisé avec option d'export Excel
            dialog = QDialog(self)
            dialog.setWindowTitle("Rapport des Dépenses")
            dialog.setMinimumSize(500, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Texte du rapport
            report_text = QTextEdit()
            report_text.setReadOnly(True)
            
            total_all = sum(e.amount for e in all_expenses)
            current_month = datetime.now().month
            month_expenses = [e for e in all_expenses if e.date.month == current_month]
            month_total = sum(e.amount for e in month_expenses)
            
            html_report = "<h3>Rapport des Dépenses</h3>"
            html_report += f"<p><b>Total dépenses:</b> {total_all:,.2f} {self.currency}</p>"
            html_report += f"<p><b>Dépenses ce mois ({datetime.now().strftime('%B')}):</b> {month_total:,.2f} {self.currency}</p>"
            html_report += "<h4>Détails par catégorie:</h4>"
            html_report += "<table border='1' style='border-collapse: collapse; width: 100%;'>"
            html_report += "<tr style='background-color: #f1f5f9;'><th>Catégorie</th><th>Nombre</th><th>Total</th><th>%</th></tr>"
            
            for category, data in sorted(categories.items(), key=lambda x: x[1]["total"], reverse=True):
                percentage = (data["total"] / total_all * 100) if total_all > 0 else 0
                html_report += f"<tr><td>{category}</td><td align='center'>{data['count']}</td><td align='right'>{data['total']:,.2f} {self.currency}</td><td align='right'>{percentage:.1f}%</td></tr>"
            
            html_report += "\\table"
            
            report_text.setHtml(html_report)
            layout.addWidget(report_text)
            
            # Boutons
            button_layout = QHBoxLayout()
            
            export_btn = QPushButton("Exporter en Excel")
            export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
            export_btn.clicked.connect(lambda: self.export_expense_report_excel(categories, total_all, month_total))
            
            close_btn = QPushButton("Fermer")
            close_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(export_btn)
            button_layout.addStretch()
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
    
    def export_expense_report_excel(self, categories, total_all, month_total):
        """Exporter le rapport des dépenses en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            
            default_name = f"rapport_depenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter le rapport des dépenses", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Feuille 1: Résumé par catégorie
                category_data = []
                for category, data in sorted(categories.items(), key=lambda x: x[1]["total"], reverse=True):
                    percentage = (data["total"] / total_all * 100) if total_all > 0 else 0
                    category_data.append({
                        'Catégorie': category,
                        'Nombre de dépenses': data['count'],
                        f'Total ({self.currency})': data['total'],
                        'Pourcentage': f"{percentage:.1f}%"
                    })
                
                df_categories = pd.DataFrame(category_data)
                df_categories.to_excel(writer, sheet_name='Par catégorie', index=False)
                
                # Feuille 2: Détail des dépenses
                all_expenses = self.db_session.query(Expense).all()
                expenses_data = []
                
                for expense in all_expenses:
                    username = expense.user.username if expense.user else "Système"
                    
                    expenses_data.append({
                        'ID': expense.id,
                        'Date': expense.date.strftime("%Y-%m-%d"),
                        'Catégorie': expense.category.name,
                        f'Montant ({self.currency})': expense.amount,
                        'Description': expense.description,
                        'Moyen de paiement': expense.payment_method or "",
                        'Référence': expense.reference or "",
                        'Utilisateur': username
                    })
                
                df_expenses = pd.DataFrame(expenses_data)
                df_expenses.to_excel(writer, sheet_name='Détail dépenses', index=False)
                
                # Feuille 3: Résumé global
                summary_data = {
                    'Statistique': [
                        'Total dépenses',
                        f'Dépenses {datetime.now().strftime("%B %Y")}',
                        'Nombre de catégories',
                        'Nombre total de dépenses'
                    ],
                    'Valeur': [
                        f"{total_all:,.2f} {self.currency}",
                        f"{month_total:,.2f} {self.currency}",
                        len(categories),
                        len(all_expenses)
                    ]
                }
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Formater automatiquement les colonnes
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Ajuster automatiquement la largeur des colonnes
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater les colonnes monétaires
                    if sheet_name == 'Par catégorie':
                        for cell in worksheet['C'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Détail dépenses':
                        for cell in worksheet['D'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"Rapport des dépenses exporté vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export: {str(e)}")
            traceback.print_exc()
    
    def export_expenses(self):
        """Exporter les dépenses en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            
            default_name = f"depenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter les dépenses", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            expenses = self.db_session.query(Expense).all()
            
            data = []
            for expense in expenses:
                username = expense.user.username if expense.user else "Système"
                
                data.append({
                    'ID': expense.id,
                    'Date': expense.date.strftime("%Y-%m-%d"),
                    'Catégorie': expense.category.name,
                    f'Montant ({self.currency})': expense.amount,
                    'Description': expense.description,
                    'Moyen de paiement': expense.payment_method or "",
                    'Référence': expense.reference or "",
                    'Utilisateur': username
                })
            
            # Exporter en Excel
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Dépenses', index=False)
                
                # Formater automatiquement les colonnes
                worksheet = writer.sheets['Dépenses']
                
                # Ajuster automatiquement la largeur des colonnes
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    col_letter = get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)
                
                # Formater les colonnes numériques
                for cell in worksheet['D'][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"{len(expenses)} dépenses exportées vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur: {str(e)}")
            traceback.print_exc()
    
    def create_reports_tab(self):
        """Créer l'onglet des rapports"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Panneau de rapports
        reports_frame = QFrame()
        reports_layout = QHBoxLayout(reports_frame)
        
        # Colonne gauche: Statistiques
        stats_col = QGroupBox("Statistiques Globales")
        stats_layout = QVBoxLayout(stats_col)
        
        self.report_total_value = QLabel(f"Valeur totale du stock: 0.00 {self.currency}")
        self.report_total_value.setObjectName("reportLabel")
        
        self.report_total_products = QLabel("Nombre de produits: 0")
        self.report_total_products.setObjectName("reportLabel")
        
        self.report_low_stock = QLabel("Produits en rupture: 0")
        self.report_low_stock.setObjectName("reportLabel")
        
        self.report_avg_margin = QLabel("Marge moyenne: 0%")
        self.report_avg_margin.setObjectName("reportLabel")
        
        self.report_total_expenses = QLabel(f"Dépenses totales: 0.00 {self.currency}")
        self.report_total_expenses.setObjectName("reportLabel")
        
        self.report_net_profit = QLabel(f"Bénéfice net estimé: 0.00 {self.currency}")
        self.report_net_profit.setObjectName("reportLabel")
        
        stats_layout.addWidget(self.report_total_value)
        stats_layout.addWidget(self.report_total_products)
        stats_layout.addWidget(self.report_low_stock)
        stats_layout.addWidget(self.report_avg_margin)
        stats_layout.addWidget(self.report_total_expenses)
        stats_layout.addWidget(self.report_net_profit)
        
        # Colonne droite: Actions
        actions_col = QGroupBox("Actions")
        actions_layout = QVBoxLayout(actions_col)
        
        btn_profit_report = QPushButton("Rapport de bénéfices")
        btn_profit_report.setObjectName("reportButton")
        btn_profit_report.setIcon(self.style().standardIcon(QStyle.SP_FileDialogInfoView))
        btn_profit_report.clicked.connect(self.generate_profit_report)
        
        btn_inventory_report = QPushButton("Rapport d'inventaire")
        btn_inventory_report.setObjectName("inventoryButton")
        btn_inventory_report.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        btn_inventory_report.clicked.connect(self.generate_inventory_report)
        
        btn_export_data = QPushButton("Exporter toutes les données")
        btn_export_data.setObjectName("exportButton")
        btn_export_data.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        btn_export_data.clicked.connect(self.export_all_data)
        
        btn_calculate_profit = QPushButton("Calculer bénéfices")
        btn_calculate_profit.setObjectName("profitButton")
        btn_calculate_profit.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
        btn_calculate_profit.clicked.connect(self.calculate_profits)
        
        for btn in [btn_profit_report, btn_inventory_report, 
                   btn_export_data, btn_calculate_profit]:
            btn.setMinimumHeight(40)
            actions_layout.addWidget(btn)
        
        reports_layout.addWidget(stats_col, 60)
        reports_layout.addWidget(actions_col, 40)
        
        layout.addWidget(reports_frame)
        
        # Tableau des produits à faible marge
        low_margin_frame = QGroupBox("Produits actifs à faible marge (< 10%)")
        low_margin_layout = QVBoxLayout(low_margin_frame)
        
        self.low_margin_table = QTableWidget()
        self.low_margin_table.setColumnCount(5)
        self.low_margin_table.setHorizontalHeaderLabels(["Code", "Produit", "Prix Achat", "Prix Vente", "Marge %"])
        
        low_margin_layout.addWidget(self.low_margin_table)
        
        layout.addWidget(low_margin_frame, 1)
        
        # Mettre à jour les statistiques
        self.update_stats()
        
        return tab
    
    def update_stats(self):
        """Mettre à jour toutes les statistiques (uniquement pour les produits actifs)"""
        try:
            # Produits actifs uniquement pour les statistiques
            active_products = [p for p in self.products if p.active]
            
            # Statistiques globales
            total_products = len(active_products)
            
            # Calculer la valeur totale du stock (uniquement actifs)
            total_value = sum(p.stock_value for p in active_products)
            
            # Produits en rupture ou stock faible (uniquement actifs)
            low_stock = len([p for p in active_products if p.is_low_stock])
            out_of_stock = len([p for p in active_products if p.is_out_of_stock])
            
            # Calculer la marge moyenne (uniquement actifs)
            margins = [p.profit_margin for p in active_products]
            avg_margin = sum(margins) / len(margins) if margins else 0
            
            # Dépenses totales
            all_expenses = self.db_session.query(Expense).all()
            total_expenses = sum(e.amount for e in all_expenses)
            
            # Bénéfice net estimé (uniquement actifs)
            estimated_profit = sum(p.total_potential_profit for p in active_products)
            net_profit = estimated_profit - total_expenses
            
            # Mettre à jour les labels de rapport
            if self.report_total_value:
                self.report_total_value.setText(f"Valeur totale du stock: {total_value:,.2f} {self.currency}")
                self.report_total_products.setText(f"Nombre de produits actifs: {total_products}")
                self.report_low_stock.setText(f"Produits en rupture/faible: {out_of_stock}/{low_stock}")
                self.report_avg_margin.setText(f"Marge moyenne: {avg_margin:.1f}%")
                self.report_total_expenses.setText(f"Dépenses totales: {total_expenses:,.2f} {self.currency}")
                self.report_net_profit.setText(f"Bénéfice net estimé: {net_profit:,.2f} {self.currency}")
            
            # Mettre à jour le tableau des faibles marges (uniquement actifs)
            if self.low_margin_table:
                low_margin_products = [p for p in active_products if p.profit_margin < 10]
                
                self.low_margin_table.setRowCount(len(low_margin_products))
                for row, product in enumerate(low_margin_products):
                    # Code
                    code_item = QTableWidgetItem(product.code or "")
                    self.low_margin_table.setItem(row, 0, code_item)
                    
                    # Nom
                    name_item = QTableWidgetItem(product.name)
                    self.low_margin_table.setItem(row, 1, name_item)
                    
                    # Prix d'achat
                    purchase_item = QTableWidgetItem(f"{product.purchase_price:,.2f} {self.currency}")
                    purchase_item.setTextAlignment(Qt.AlignRight)
                    self.low_margin_table.setItem(row, 2, purchase_item)
                    
                    # Prix de vente
                    sale_item = QTableWidgetItem(f"{product.sale_price:,.2f} {self.currency}")
                    sale_item.setTextAlignment(Qt.AlignRight)
                    self.low_margin_table.setItem(row, 3, sale_item)
                    
                    # Marge
                    margin_item = QTableWidgetItem(f"{product.profit_margin:.1f}%")
                    margin_item.setTextAlignment(Qt.AlignCenter)
                    margin_item.setForeground(QColor("#ef4444"))
                    
                    self.low_margin_table.setItem(row, 4, margin_item)
                
                # Ajuster les colonnes
                self.low_margin_table.resizeColumnsToContents()
            
        except Exception as e:
            print(f"Erreur lors de la mise à jour des statistiques: {e}")
    
    def generate_profit_report(self):
        """Générer un rapport de bénéfices"""
        try:
            # Utiliser uniquement les produits actifs
            active_products = [p for p in self.products if p.active]
            
            # Calculer les totaux
            total_investment = sum(p.stock_value for p in active_products)
            total_potential = sum(p.quantity * p.sale_price for p in active_products)
            total_profit = total_potential - total_investment
            
            # Dépenses
            all_expenses = self.db_session.query(Expense).all()
            total_expenses = sum(e.amount for e in all_expenses)
            net_profit = total_profit - total_expenses
            
            # Catégories de marge
            high_margin = [p for p in active_products if p.profit_margin > 30]
            medium_margin = [p for p in active_products if 10 <= p.profit_margin <= 30]
            low_margin = [p for p in active_products if p.profit_margin < 10]
            
            # Créer un dialogue personnalisé
            dialog = QDialog(self)
            dialog.setWindowTitle("Rapport de Bénéfices")
            dialog.setMinimumSize(500, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Texte du rapport
            report_text = QTextEdit()
            report_text.setReadOnly(True)
            report_text.setHtml(f"""
            <h3 style='text-align: center;'>Rapport de Bénéfices</h3>
            
            <table style='border-collapse: collapse; width: 100%; margin: 20px 0;'>
                <tr style='background-color: #f8fafc;'>
                    <td style='padding: 12px; font-weight: bold; border: 1px solid #e2e8f0;'>Investissement total en stock:</td>
                    <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; font-weight: bold;'>{total_investment:,.2f} {self.currency}</td>
                </tr>
                <tr>
                    <td style='padding: 12px; border: 1px solid #e2e8f0;'>Valeur potentielle de vente:</td>
                    <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0;'>{total_potential:,.2f} {self.currency}</td>
                </tr>
                <tr style='background-color: #f0fdf4;'>
                    <td style='padding: 12px; font-weight: bold; border: 1px solid #e2e8f0; color: #15803d;'>Profit brut potentiel:</td>
                    <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; font-weight: bold; color: #15803d;'>{total_profit:,.2f} {self.currency}</td>
                </tr>
                <tr>
                    <td style='padding: 12px; border: 1px solid #e2e8f0;'>Dépenses totales enregistrées:</td>
                    <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; color: #dc2626;'>{total_expenses:,.2f} {self.currency}</td>
                </tr>
                <tr style='background-color: #dbeafe; border-top: 3px solid #3b82f6;'>
                    <td style='padding: 15px; font-weight: bold; font-size: 16px; border: 1px solid #e2e8f0;'>BÉNÉFICE NET ESTIMÉ:</td>
                    <td style='padding: 15px; text-align: right; font-weight: bold; font-size: 16px; color: #1d4ed8; border: 1px solid #e2e8f0;'>
                        {net_profit:,.2f} {self.currency}
                    </td>
                </tr>
            </table>
            
            <div style='margin: 20px 0; padding: 15px; background-color: #f8fafc; border-radius: 8px;'>
                <p><b>Taux de marge:</b> {(total_profit/total_investment*100) if total_investment > 0 else 0:.1f}%</p>
                <p><b>Retour sur investissement (ROI):</b> {(net_profit/total_investment*100) if total_investment > 0 else 0:.1f}%</p>
                <p><b>Marge bénéficiaire nette:</b> {(net_profit/total_potential*100) if total_potential > 0 else 0:.1f}%</p>
            </div>
            
            <h4 style='margin-top: 20px;'>Analyse par marge:</h4>
            <ul>
                <li>Produits à haute marge (> 30%): {len(high_margin):,}</li>
                <li>Produits à marge moyenne (10-30%): {len(medium_margin):,}</li>
                <li>Produits à faible marge (< 10%): {len(low_margin):,}</li>
            </ul>
            """)
            
            layout.addWidget(report_text)
            
            # Boutons
            button_layout = QHBoxLayout()
            
            export_btn = QPushButton("Exporter en Excel")
            export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
            export_btn.clicked.connect(lambda: self.export_profit_report_excel(
                total_investment, total_potential, total_profit, 
                total_expenses, net_profit, high_margin, medium_margin, low_margin
            ))
            
            close_btn = QPushButton("Fermer")
            close_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(export_btn)
            button_layout.addStretch()
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
    
    def export_profit_report_excel(self, total_investment, total_potential, total_profit, 
                                   total_expenses, net_profit, high_margin, medium_margin, low_margin):
        """Exporter le rapport de bénéfices en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            
            default_name = f"rapport_benefices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter le rapport de bénéfices", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Feuille 1: Résumé
                summary_data = {
                    'Élément': [
                        'Investissement total en stock',
                        'Valeur potentielle de vente',
                        'Profit brut potentiel',
                        'Dépenses totales',
                        'BÉNÉFICE NET ESTIMÉ'
                    ],
                    f'Montant ({self.currency})': [
                        total_investment,
                        total_potential,
                        total_profit,
                        total_expenses,
                        net_profit
                    ],
                    'Pourcentage': [
                        '100.0%',
                        f'{(total_potential/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(total_profit/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(total_expenses/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(net_profit/total_investment*100) if total_investment > 0 else 0:.1f}%'
                    ]
                }
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Feuille 2: Analyse par marge
                active_products = [p for p in self.products if p.active]
                margin_data = {
                    'Catégorie': [
                        'Haute marge (> 30%)',
                        'Marge moyenne (10-30%)',
                        'Faible marge (< 10%)',
                        'TOTAL'
                    ],
                    'Nombre de produits': [
                        len(high_margin),
                        len(medium_margin),
                        len(low_margin),
                        len(active_products)
                    ],
                    f'Valeur du stock ({self.currency})': [
                        sum(p.stock_value for p in high_margin),
                        sum(p.stock_value for p in medium_margin),
                        sum(p.stock_value for p in low_margin),
                        total_investment
                    ],
                    'Pourcentage': [
                        f'{(sum(p.stock_value for p in high_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(sum(p.stock_value for p in medium_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(sum(p.stock_value for p in low_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        '100.0%'
                    ]
                }
                
                df_margin = pd.DataFrame(margin_data)
                df_margin.to_excel(writer, sheet_name='Analyse par marge', index=False)
                
                # Feuille 3: Détail des produits actifs
                products_data = []
                for product in active_products:
                    supplier_name = product.supplier.name if product.supplier else ""
                    margin_category = "Haute" if product.profit_margin > 30 else "Moyenne" if product.profit_margin >= 10 else "Faible"
                    
                    products_data.append({
                        'Code': product.code or "",
                        'Nom': product.name,
                        'Catégorie': product.category,
                        'Quantité': product.quantity,
                        'Prix Achat': product.purchase_price,
                        'Prix Vente': product.sale_price,
                        f'Valeur Stock ({self.currency})': product.stock_value,
                        'Marge %': f"{product.profit_margin:.1f}%",
                        'Catégorie Marge': margin_category,
                        'Fournisseur': supplier_name,
                        'Actif': "Oui" if product.active else "Non"
                    })
                
                df_products = pd.DataFrame(products_data)
                df_products.to_excel(writer, sheet_name='Détail Produits', index=False)
                
                # Formater automatiquement les colonnes
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Ajuster automatiquement la largeur des colonnes
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater les colonnes monétaires
                    if sheet_name == 'Résumé':
                        for cell in worksheet['B'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Analyse par marge':
                        for cell in worksheet['C'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Détail Produits':
                        for cell in worksheet['F'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['H'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"Rapport de bénéfices exporté vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export: {str(e)}")
            traceback.print_exc()
    
    def generate_inventory_report(self):
        """Générer un rapport d'inventaire"""
        try:
            # Utiliser tous les produits (actifs et inactifs) pour l'inventaire complet
            # Inventaire par catégorie
            categories = {}
            for product in self.products:
                if product.category not in categories:
                    categories[product.category] = {"count": 0, "value": 0, "quantity": 0}
                
                categories[product.category]["count"] += 1
                categories[product.category]["value"] += product.stock_value
                categories[product.category]["quantity"] += product.quantity
            
            # Produits en rupture (actifs uniquement)
            active_products = [p for p in self.products if p.active]
            low_stock_products = [p for p in active_products if p.is_low_stock]
            out_of_stock_products = [p for p in active_products if p.is_out_of_stock]
            
            # Statistiques globales
            total_value = sum(p.stock_value for p in self.products)
            total_quantity = sum(p.quantity for p in self.products)
            
            # Créer un dialogue personnalisé
            dialog = QDialog(self)
            dialog.setWindowTitle("Rapport d'Inventaire")
            dialog.setMinimumSize(500, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Texte du rapport
            report_text = QTextEdit()
            report_text.setReadOnly(True)
            
            html_report = "<h3>Rapport d'Inventaire</h3>"
            html_report += f"<p><b>Valeur totale du stock:</b> {total_value:,.2f} {self.currency}</p>"
            html_report += f"<p><b>Quantité totale en stock:</b> {total_quantity:,} unités</p>"
            
            html_report += "<h4>Détails par catégorie:</h4>"
            html_report += "<table border='1' style='border-collapse: collapse; width: 100%;'>"
            html_report += "<tr><th>Catégorie</th><th>Produits</th><th>Quantité</th><th>Valeur</th><th>%</th></tr>"
            
            for category, data in sorted(categories.items(), key=lambda x: x[1]["value"], reverse=True):
                percentage = (data["value"] / total_value * 100) if total_value > 0 else 0
                html_report += f"<tr><td>{category}</td><td align='center'>{data['count']}</td><td align='right'>{data['quantity']:,}</td><td align='right'>{data['value']:,.2f} {self.currency}</td><td align='right'>{percentage:.1f}%</td></tr>"
            
            html_report += "\\table"
            
            # Produits en rupture
            if out_of_stock_products:
                html_report += "<h4 style='color: #dc2626;'>Produits actifs en rupture:</h4>"
                html_report += "<ul>"
                for product in out_of_stock_products:
                    html_report += f"<li>{product.name} (Code: {product.code or 'N/A'})</li>"
                html_report += "</ul>"
            
            if low_stock_products:
                html_report += "<h4 style='color: #f59e0b;'>Produits actifs en stock faible:</h4>"
                html_report += "<ul>"
                for product in low_stock_products:
                    html_report += f"<li>{product.name} (Stock: {product.quantity}, Minimum: {product.min_stock})</li>"
                html_report += "</ul>"
            
            report_text.setHtml(html_report)
            layout.addWidget(report_text)
            
            # Boutons
            button_layout = QHBoxLayout()
            
            export_btn = QPushButton("Exporter en Excel")
            export_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
            export_btn.clicked.connect(lambda: self.export_inventory_report_excel(
                categories, total_value, total_quantity, low_stock_products, out_of_stock_products
            ))
            
            close_btn = QPushButton("Fermer")
            close_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(export_btn)
            button_layout.addStretch()
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
    
    def export_inventory_report_excel(self, categories, total_value, total_quantity, low_stock_products, out_of_stock_products):
        """Exporter le rapport d'inventaire en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            
            default_name = f"rapport_inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter le rapport d'inventaire", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Feuille 1: Résumé par catégorie
                category_data = []
                for category, data in sorted(categories.items(), key=lambda x: x[1]["value"], reverse=True):
                    percentage = (data["value"] / total_value * 100) if total_value > 0 else 0
                    category_data.append({
                        'Catégorie': category,
                        'Nombre de produits': data['count'],
                        'Quantité totale': data['quantity'],
                        f'Valeur ({self.currency})': data['value'],
                        'Pourcentage': f"{percentage:.1f}%"
                    })
                
                df_categories = pd.DataFrame(category_data)
                df_categories.to_excel(writer, sheet_name='Par catégorie', index=False)
                
                # Feuille 2: Produits en rupture/stock faible (actifs uniquement)
                problematic_products = []
                for product in out_of_stock_products:
                    problematic_products.append({
                        'Type': 'RUPTURE',
                        'Code': product.code or "",
                        'Nom': product.name,
                        'Catégorie': product.category,
                        'Stock actuel': product.quantity,
                        'Stock minimum': product.min_stock,
                        f'Valeur ({self.currency})': product.stock_value,
                        'Statut': 'Actif'
                    })
                
                for product in low_stock_products:
                    if product not in out_of_stock_products:
                        problematic_products.append({
                            'Type': 'STOCK FAIBLE',
                            'Code': product.code or "",
                            'Nom': product.name,
                            'Catégorie': product.category,
                            'Stock actuel': product.quantity,
                            'Stock minimum': product.min_stock,
                            f'Valeur ({self.currency})': product.stock_value,
                            'Statut': 'Actif'
                        })
                
                df_problematic = pd.DataFrame(problematic_products)
                df_problematic.to_excel(writer, sheet_name='Produits problématiques', index=False)
                
                # Feuille 3: Détail de tous les produits
                products_data = []
                for product in self.products:
                    supplier_name = product.supplier.name if product.supplier else ""
                    stock_status = "Normal"
                    if product.active:
                        if product.is_out_of_stock:
                            stock_status = "Rupture"
                        elif product.is_low_stock:
                            stock_status = "Faible"
                    else:
                        stock_status = "Inactif"
                    
                    products_data.append({
                        'Code': product.code or "",
                        'Nom': product.name,
                        'Catégorie': product.category,
                        'Quantité': product.quantity,
                        'Stock minimum': product.min_stock,
                        'Stock maximum': product.max_stock,
                        f'Valeur ({self.currency})': product.stock_value,
                        'Statut stock': stock_status,
                        'Statut produit': 'Actif' if product.active else 'Inactif',
                        'Fournisseur': supplier_name
                    })
                
                df_products = pd.DataFrame(products_data)
                df_products.to_excel(writer, sheet_name='Détail produits', index=False)
                
                # Feuille 4: Résumé global
                inactive_count = len([p for p in self.products if not p.active])
                summary_data = {
                    'Statistique': [
                        'Valeur totale du stock',
                        'Quantité totale en stock',
                        'Nombre total de produits',
                        'Produits actifs',
                        'Produits inactifs',
                        'Nombre de catégories',
                        'Produits actifs en rupture',
                        'Produits actifs en stock faible',
                        'Produits actifs avec stock normal'
                    ],
                    'Valeur': [
                        f"{total_value:,.2f} {self.currency}",
                        total_quantity,
                        len(self.products),
                        len([p for p in self.products if p.active]),
                        inactive_count,
                        len(categories),
                        len(out_of_stock_products),
                        len(low_stock_products),
                        len([p for p in self.products if p.active and not p.is_low_stock and not p.is_out_of_stock])
                    ]
                }
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Formater automatiquement les colonnes
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Ajuster automatiquement la largeur des colonnes
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater les colonnes monétaires
                    if sheet_name == 'Par catégorie':
                        for cell in worksheet['D'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Produits problématiques':
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Détail produits':
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"Rapport d'inventaire exporté vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export: {str(e)}")
            traceback.print_exc()
    
    def calculate_profits(self):
        """Calculer les bénéfices détaillés"""
        dialog = ProfitCalculationDialog(self, [p for p in self.products if p.active], self.db_session)
        dialog.exec()
    
    def export_all_data(self):
        """Exporter toutes les données"""
        format_dialog = QInputDialog()
        format_dialog.setWindowTitle("Format d'export")
        format_dialog.setLabelText("Choisissez le format d'export:")
        format_dialog.setComboBoxItems(["Excel (XLSX)", "JSON"])
        format_dialog.setComboBoxEditable(False)
        
        if format_dialog.exec():
            export_format = format_dialog.textValue()
            
            try:
                if export_format == "Excel (XLSX)":
                    self.export_to_excel()
                elif export_format == "JSON":
                    self.export_to_json()
                    
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export: {str(e)}")
    
    def export_to_excel(self):
        """Exporter toutes les données en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            
            default_name = f"stock_complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter toutes les données", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            # Créer un classeur Excel avec plusieurs feuilles
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Feuille 1: Produits
                products_data = []
                for product in self.products:
                    supplier_name = product.supplier.name if product.supplier else ""
                    
                    products_data.append({
                        'ID': product.id,
                        'Code': product.code or "",
                        'Nom': product.name,
                        'Catégorie': product.category,
                        'Quantité': product.quantity,
                        'Prix Achat': product.purchase_price,
                        'Prix Vente': product.sale_price,
                        'Fournisseur': supplier_name,
                        'Stock Min': product.min_stock,
                        'Stock Max': product.max_stock,
                        f'Valeur Stock ({self.currency})': product.stock_value,
                        'Marge %': f"{product.profit_margin:.1f}%",
                        'Actif': "Oui" if product.active else "Non"
                    })
                
                df_products = pd.DataFrame(products_data)
                df_products.to_excel(writer, sheet_name='Produits', index=False)
                
                # Feuille 2: Mouvements
                movements = self.db_session.query(InventoryMovement).all()
                movements_data = []
                
                for movement in movements:
                    product_name = movement.product.name if movement.product else f"ID:{movement.product_id}"
                    username = movement.user.username if movement.user else "Système"
                    
                    movements_data.append({
                        'ID': movement.id,
                        'Date': movement.date.strftime("%Y-%m-%d %H:%M"),
                        'Produit': product_name,
                        'Type': movement.movement_type,
                        'Quantité': movement.quantity,
                        'Prix unitaire': movement.unit_price or 0,
                        f'Valeur totale ({self.currency})': movement.total_value or 0,
                        'Raison': movement.reason or "",
                        'Utilisateur': username,
                        'Notes': movement.notes or ""
                    })
                
                df_movements = pd.DataFrame(movements_data)
                df_movements.to_excel(writer, sheet_name='Mouvements', index=False)
                
                # Feuille 3: Dépenses
                expenses = self.db_session.query(Expense).all()
                expenses_data = []
                
                for expense in expenses:
                    username = expense.user.username if expense.user else "Système"
                    
                    expenses_data.append({
                        'ID': expense.id,
                        'Date': expense.date.strftime("%Y-%m-%d"),
                        'Catégorie': expense.category.name,
                        f'Montant ({self.currency})': expense.amount,
                        'Description': expense.description,
                        'Moyen de paiement': expense.payment_method or "",
                        'Référence': expense.reference or "",
                        'Utilisateur': username
                    })
                
                df_expenses = pd.DataFrame(expenses_data)
                df_expenses.to_excel(writer, sheet_name='Dépenses', index=False)
                
                # Feuille 4: Ventes
                sales = self.db_session.query(Sale).all()
                sales_data = []
                
                for sale in sales:
                    customer_name = ""
                    if sale.customer:
                        customer_name = f"{sale.customer.first_name or ''} {sale.customer.last_name or ''}".strip()
                        if not customer_name and sale.customer.company:
                            customer_name = sale.customer.company
                    
                    cashier_name = sale.cashier.username if sale.cashier else "Inconnu"
                    
                    sales_data.append({
                        'N° Vente': sale.sale_number,
                        'Date': sale.sale_date.strftime("%Y-%m-%d %H:%M"),
                        'Client': customer_name,
                        f'Sous-total ({self.currency})': sale.subtotal,
                        f'Remise ({self.currency})': sale.discount_amount,
                        f'TVA ({self.currency})': sale.tax_amount,
                        f'Total ({self.currency})': sale.total_amount,
                        'Paiement': sale.payment_method,
                        'Statut': sale.sale_status,
                        'Caissier': cashier_name
                    })
                
                df_sales = pd.DataFrame(sales_data)
                df_sales.to_excel(writer, sheet_name='Ventes', index=False)
                
                # Feuille 5: Résumé
                active_products = [p for p in self.products if p.active]
                total_expenses = sum(e.amount for e in expenses)
                total_sales = sum(s.total_amount for s in sales)
                
                summary_data = {
                    'Statistique': [
                        'Total produits (tous)',
                        'Produits actifs',
                        'Produits inactifs',
                        f'Valeur totale du stock ({self.currency})',
                        'Quantité totale en stock',
                        'Produits en rupture (actifs)',
                        'Produits en stock faible (actifs)',
                        'Marge moyenne (actifs)',
                        f'Dépenses totales ({self.currency})',
                        f'Ventes totales ({self.currency})',
                        f'Bénéfice net estimé ({self.currency})'
                    ],
                    'Valeur': [
                        len(self.products),
                        len(active_products),
                        len([p for p in self.products if not p.active]),
                        f"{sum(p.stock_value for p in self.products):,.2f}",
                        sum(p.quantity for p in self.products),
                        len([p for p in active_products if p.is_out_of_stock]),
                        len([p for p in active_products if p.is_low_stock]),
                        f"{sum(p.profit_margin for p in active_products) / len(active_products) if active_products else 0:.1f}%",
                        f"{total_expenses:,.2f}",
                        f"{total_sales:,.2f}",
                        f"{total_sales - total_expenses:,.2f}"
                    ]
                }
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Formater automatiquement toutes les feuilles
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Ajuster automatiquement la largeur des colonnes
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater les colonnes monétaires selon la feuille
                    if sheet_name == 'Produits':
                        for cell in worksheet['F'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['K'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Mouvements':
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['H'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Dépenses':
                        for cell in worksheet['D'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Ventes':
                        for cell in worksheet['D'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['E'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['F'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", 
                                  f"Toutes les données exportées vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export Excel: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def export_to_json(self):
        """Exporter en JSON"""
        from datetime import datetime
        
        # Demander où sauvegarder
        default_name = f"stock_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filename, _ = QFileDialog.getSaveFileName(
            self, "Exporter en JSON", default_name,
            "Fichiers JSON (*.json);;Tous les fichiers (*.*)"
        )
        
        if not filename:
            return
        
        # Préparer les données
        data = {
            "products": [
                {
                    "id": p.id,
                    "code": p.code,
                    "name": p.name,
                    "category": p.category,
                    "quantity": p.quantity,
                    "purchase_price": p.purchase_price,
                    "sale_price": p.sale_price,
                    "supplier": p.supplier.name if p.supplier else None,
                    "min_stock": p.min_stock,
                    "max_stock": p.max_stock,
                    "location": p.location,
                    "active": p.active,
                    "stock_value": p.stock_value,
                    "profit_margin": p.profit_margin
                } for p in self.products
            ],
            "movements": [
                {
                    "id": m.id,
                    "product_id": m.product_id,
                    "product_name": m.product.name if m.product else None,
                    "movement_type": m.movement_type,
                    "quantity": m.quantity,
                    "unit_price": m.unit_price,
                    "total_value": m.total_value,
                    "reason": m.reason,
                    "date": m.date.isoformat(),
                    "user": m.user.username if m.user else None
                } for m in self.db_session.query(InventoryMovement).all()
            ],
            "expenses": [
                {
                    "id": e.id,
                    "category": e.category.name,
                    "amount": e.amount,
                    "description": e.description,
                    "payment_method": e.payment_method,
                    "date": e.date.isoformat(),
                    "user": e.user.username if e.user else None
                } for e in self.db_session.query(Expense).all()
            ],
            "sales": [
                {
                    "id": s.id,
                    "sale_number": s.sale_number,
                    "sale_date": s.sale_date.isoformat() if s.sale_date else None,
                    "customer_id": s.customer_id,
                    "subtotal": s.subtotal,
                    "discount_amount": s.discount_amount,
                    "tax_amount": s.tax_amount,
                    "total_amount": s.total_amount,
                    "payment_method": s.payment_method,
                    "sale_status": s.sale_status,
                    "cashier_id": s.cashier_id
                } for s in self.db_session.query(Sale).all()
            ],
            "export_date": datetime.now().isoformat(),
            "export_user": self.user.get('username', 'System') if isinstance(self.user, dict) else self.user.username,
            "company_info": self.company_info,
            "currency": self.currency,
            "tax_rate": self.tax_rate
        }
        
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        QMessageBox.information(self, "Succès", f"Données exportées en JSON vers:\n{filename}")
    
    def refresh_data(self):
        """Actualiser les données"""
        try:
            self.load_data()
            self.filter_products()
            self.load_movements_table()
            self.load_expenses_table()
            self.load_sales_table()
            self.update_stats()
            
            QMessageBox.information(self, "Actualisation", "Données actualisées avec succès!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'actualisation: {str(e)}")
    
    def closeEvent(self, event):
        """Fermer la session de base de données"""
        self.db_session.close()
        super().closeEvent(event)





# ===== CLASSES DE DIALOGUES =====

class ProductDialog(QDialog):
    """Boîte de dialogue pour ajouter/modifier un produit"""
    def __init__(self, parent=None, db_session=None, product=None):
        super().__init__(parent)
        
        self.db_session = db_session
        self.product = product
        
        self.setWindowTitle("Nouveau Produit" if not product else "Modifier Produit")
        self.setModal(True)
        self.setMinimumWidth(600)
        
        layout = QVBoxLayout(self)
        
        # Formulaire
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        
        # Code produit
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("Ex: PROD001")
        
        # Nom produit
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Nom du produit")
        
        # Catégorie
        self.category_input = QLineEdit()
        self.category_input.setPlaceholderText("Ex: Informatique, Alimentaire, etc.")
        
        # Description
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(80)
        self.description_input.setPlaceholderText("Description du produit")
        
        # Quantité
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(0, 99999)
        self.quantity_spin.setValue(0)
        
        # Prix d'achat
        self.purchase_spin = QDoubleSpinBox()
        self.purchase_spin.setRange(0, 99999999.99)
        self.purchase_spin.setDecimals(2)
        self.purchase_spin.setPrefix("FCFA ")
        self.purchase_spin.setValue(0)
        
        # Prix de vente
        self.sale_spin = QDoubleSpinBox()
        self.sale_spin.setRange(0, 99999999.99)
        self.sale_spin.setDecimals(2)
        self.sale_spin.setPrefix("FCFA ")
        self.sale_spin.setValue(0)
        
        # Fournisseur
        self.supplier_combo = QComboBox()
        self.supplier_combo.addItem("-- Sélectionner --", None)
        
        # Charger les fournisseurs
        suppliers = db_session.query(Supplier).filter(Supplier.active == True).all()
        for supplier in suppliers:
            self.supplier_combo.addItem(supplier.name, supplier.id)
        
        # Bouton Nouveau fournisseur
        self.new_supplier_btn = QPushButton("➕ Nouveau")
        self.new_supplier_btn.setObjectName("smallButton")
        self.new_supplier_btn.clicked.connect(self.show_new_supplier_dialog)
        
        supplier_layout = QHBoxLayout()
        supplier_layout.addWidget(self.supplier_combo, 80)
        supplier_layout.addWidget(self.new_supplier_btn, 20)
        
        # Stock minimum
        self.min_stock_spin = QSpinBox()
        self.min_stock_spin.setRange(0, 9999)
        self.min_stock_spin.setValue(5)
        
        # Stock maximum
        self.max_stock_spin = QSpinBox()
        self.max_stock_spin.setRange(0, 99999)
        self.max_stock_spin.setValue(100)
        
        # Emplacement
        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("Ex: Rayon A, Étagère 3")
        
        # Code-barres
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("Code-barres (optionnel)")
        
        # Statut
        self.active_checkbox = QCheckBox("Actif")
        self.active_checkbox.setChecked(True)
        
        # Ajouter les champs au formulaire
        form.addRow("Code produit*:", self.code_input)
        form.addRow("Nom*:", self.name_input)
        form.addRow("Catégorie*:", self.category_input)
        form.addRow("Description:", self.description_input)
        form.addRow("Quantité initiale:", self.quantity_spin)
        form.addRow("Prix d'achat*:", self.purchase_spin)
        form.addRow("Prix de vente*:", self.sale_spin)
        form.addRow("Fournisseur:", supplier_layout)
        form.addRow("Stock minimum:", self.min_stock_spin)
        form.addRow("Stock maximum:", self.max_stock_spin)
        form.addRow("Emplacement:", self.location_input)
        form.addRow("Code-barres:", self.barcode_input)
        form.addRow("Statut:", self.active_checkbox)
        
        layout.addLayout(form)
        
        # Calcul de marge en temps réel
        self.margin_label = QLabel("Marge: 0.00%")
        self.margin_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.margin_label)
        
        # Connecter les changements de prix
        self.purchase_spin.valueChanged.connect(self.update_margin)
        self.sale_spin.valueChanged.connect(self.update_margin)
        
        # Pré-remplir si modification
        if product:
            self.code_input.setText(product.code or "")
            self.name_input.setText(product.name)
            self.category_input.setText(product.category)
            self.description_input.setPlainText(product.description or "")
            self.quantity_spin.setValue(product.quantity)
            self.purchase_spin.setValue(product.purchase_price)
            self.sale_spin.setValue(product.sale_price)
            
            if product.supplier_id:
                for i in range(self.supplier_combo.count()):
                    if self.supplier_combo.itemData(i) == product.supplier_id:
                        self.supplier_combo.setCurrentIndex(i)
                        break
            
            self.min_stock_spin.setValue(product.min_stock)
            self.max_stock_spin.setValue(product.max_stock)
            self.location_input.setText(product.location or "")
            self.barcode_input.setText(product.barcode or "")
            self.active_checkbox.setChecked(product.active)
            
            self.update_margin()
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def show_new_supplier_dialog(self):
        """Afficher la boîte de dialogue pour créer un nouveau fournisseur"""
        dialog = SupplierDialog(self, self.db_session)
        if dialog.exec():
            # Recharger les fournisseurs dans la combobox
            self.supplier_combo.clear()
            self.supplier_combo.addItem("-- Sélectionner --", None)
            
            suppliers = self.db_session.query(Supplier).filter(Supplier.active == True).all()
            for supplier in suppliers:
                self.supplier_combo.addItem(supplier.name, supplier.id)
            
            # Sélectionner le nouveau fournisseur
            self.supplier_combo.setCurrentText(dialog.name_input.text())
            
            QMessageBox.information(self, "Succès", "Fournisseur ajouté avec succès!")
    
    def update_margin(self):
        """Mettre à jour l'affichage de la marge"""
        purchase = self.purchase_spin.value()
        sale = self.sale_spin.value()
        
        if purchase > 0:
            margin = ((sale - purchase) / purchase * 100)
            self.margin_label.setText(f"Marge: {margin:.2f}%")
            
            if margin < 10:
                self.margin_label.setStyleSheet("color: #ef4444; font-weight: bold;")
            elif margin > 30:
                self.margin_label.setStyleSheet("color: #10b981; font-weight: bold;")
            else:
                self.margin_label.setStyleSheet("color: #000000; font-weight: bold;")
        else:
            self.margin_label.setText("Marge: N/A")
            self.margin_label.setStyleSheet("color: #666; font-weight: bold;")
    
    def validate_and_save(self):
        """Valider et sauvegarder le produit"""
        # Validation
        if not self.code_input.text().strip():
            QMessageBox.warning(self, "Validation", "Le code produit est obligatoire!")
            return
        
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Validation", "Le nom du produit est obligatoire!")
            return
        
        if not self.category_input.text().strip():
            QMessageBox.warning(self, "Validation", "La catégorie est obligatoire!")
            return
        
        if self.purchase_spin.value() <= 0:
            QMessageBox.warning(self, "Validation", "Le prix d'achat doit être supérieur à 0!")
            return
        
        if self.sale_spin.value() <= 0:
            QMessageBox.warning(self, "Validation", "Le prix de vente doit être supérieur à 0!")
            return
        
        # Vérifier l'unicité du code
        existing = self.db_session.query(Product).filter(
            Product.code == self.code_input.text().strip()
        ).first()
        
        if existing and (not self.product or existing.id != self.product.id):
            QMessageBox.warning(self, "Validation", 
                              f"Le code '{self.code_input.text()}' est déjà utilisé!")
            return
        
        # Sauvegarder
        try:
            if self.product:
                # Modification
                self.product.code = self.code_input.text().strip()
                self.product.name = self.name_input.text().strip()
                self.product.category = self.category_input.text().strip()
                self.product.description = self.description_input.toPlainText().strip() or None
                self.product.quantity = self.quantity_spin.value()
                self.product.purchase_price = self.purchase_spin.value()
                self.product.sale_price = self.sale_spin.value()
                self.product.supplier_id = self.supplier_combo.currentData()
                self.product.min_stock = self.min_stock_spin.value()
                self.product.max_stock = self.max_stock_spin.value()
                self.product.location = self.location_input.text().strip() or None
                self.product.barcode = self.barcode_input.text().strip() or None
                self.product.active = self.active_checkbox.isChecked()
            else:
                # Nouveau produit
                new_product = Product(
                    code=self.code_input.text().strip(),
                    name=self.name_input.text().strip(),
                    category=self.category_input.text().strip(),
                    description=self.description_input.toPlainText().strip() or None,
                    quantity=self.quantity_spin.value(),
                    purchase_price=self.purchase_spin.value(),
                    sale_price=self.sale_spin.value(),
                    supplier_id=self.supplier_combo.currentData(),
                    min_stock=self.min_stock_spin.value(),
                    max_stock=self.max_stock_spin.value(),
                    location=self.location_input.text().strip() or None,
                    barcode=self.barcode_input.text().strip() or None,
                    active=self.active_checkbox.isChecked()
                )
                
                self.db_session.add(new_product)
            
            self.db_session.commit()
            self.accept()
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde: {str(e)}")


class SupplierDialog(QDialog):
    """Boîte de dialogue pour ajouter un nouveau fournisseur"""
    def __init__(self, parent=None, db_session=None, supplier=None):
        super().__init__(parent)
        
        self.db_session = db_session
        self.supplier = supplier
        
        self.setWindowTitle("Nouveau Fournisseur" if not supplier else "Modifier Fournisseur")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        # Formulaire
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        
        # Nom
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Nom du fournisseur")
        
        # Code
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("Code fournisseur (optionnel)")
        
        # Contact
        self.contact_input = QLineEdit()
        self.contact_input.setPlaceholderText("Personne à contacter")
        
        # Téléphone
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Téléphone")
        
        # Email
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Email")
        
        # Adresse
        self.address_input = QTextEdit()
        self.address_input.setMaximumHeight(80)
        self.address_input.setPlaceholderText("Adresse complète")
        
        # Site web
        self.website_input = QLineEdit()
        self.website_input.setPlaceholderText("Site web")
        
        # Notes
        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(80)
        self.notes_input.setPlaceholderText("Notes supplémentaires")
        
        # Statut
        self.active_checkbox = QCheckBox("Actif")
        self.active_checkbox.setChecked(True)
        
        # Ajouter les champs au formulaire
        form.addRow("Nom*:", self.name_input)
        form.addRow("Code:", self.code_input)
        form.addRow("Contact:", self.contact_input)
        form.addRow("Téléphone:", self.phone_input)
        form.addRow("Email:", self.email_input)
        form.addRow("Adresse:", self.address_input)
        form.addRow("Site web:", self.website_input)
        form.addRow("Notes:", self.notes_input)
        form.addRow("Statut:", self.active_checkbox)
        
        layout.addLayout(form)
        
        # Pré-remplir si modification
        if supplier:
            self.name_input.setText(supplier.name)
            self.code_input.setText(supplier.code or "")
            self.contact_input.setText(supplier.contact_person or "")
            self.phone_input.setText(supplier.phone or "")
            self.email_input.setText(supplier.email or "")
            self.address_input.setPlainText(supplier.address or "")
            self.website_input.setText(supplier.website or "")
            self.notes_input.setPlainText(supplier.notes or "")
            self.active_checkbox.setChecked(supplier.active)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def validate_and_save(self):
        """Valider et sauvegarder le fournisseur"""
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Validation", "Le nom du fournisseur est obligatoire!")
            return
        
        try:
            if self.supplier:
                self.supplier.name = self.name_input.text().strip()
                self.supplier.code = self.code_input.text().strip() or None
                self.supplier.contact_person = self.contact_input.text().strip() or None
                self.supplier.phone = self.phone_input.text().strip() or None
                self.supplier.email = self.email_input.text().strip() or None
                self.supplier.address = self.address_input.toPlainText().strip() or None
                self.supplier.website = self.website_input.text().strip() or None
                self.supplier.notes = self.notes_input.toPlainText().strip() or None
                self.supplier.active = self.active_checkbox.isChecked()
            else:
                new_supplier = Supplier(
                    name=self.name_input.text().strip(),
                    code=self.code_input.text().strip() or None,
                    contact_person=self.contact_input.text().strip() or None,
                    phone=self.phone_input.text().strip() or None,
                    email=self.email_input.text().strip() or None,
                    address=self.address_input.toPlainText().strip() or None,
                    website=self.website_input.text().strip() or None,
                    notes=self.notes_input.toPlainText().strip() or None,
                    active=self.active_checkbox.isChecked()
                )
                self.db_session.add(new_supplier)
            
            self.db_session.commit()
            self.accept()
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde: {str(e)}")


class StockOperationDialog(QDialog):
    """Boîte de dialogue pour les opérations de stock"""
    def __init__(self, parent=None, product=None, operation_type="IN"):
        super().__init__(parent)
        
        self.product = product
        self.operation_type = operation_type
        
        self.setWindowTitle(f"{'Entrée' if operation_type == 'IN' else 'Sortie'} de stock")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Informations produit
        info_group = QGroupBox("Informations du produit")
        info_layout = QVBoxLayout(info_group)
        
        if product:
            info_layout.addWidget(QLabel(f"<b>Produit:</b> {product.name}"))
            info_layout.addWidget(QLabel(f"<b>Code:</b> {product.code or 'N/A'}"))
            info_layout.addWidget(QLabel(f"<b>Stock actuel:</b> {product.quantity:,} unités"))
            
            if operation_type == "OUT" and product.quantity <= 0:
                warning_label = QLabel("⚠️ Ce produit est en rupture de stock!")
                warning_label.setStyleSheet("color: #ef4444; font-weight: bold;")
                info_layout.addWidget(warning_label)
        
        layout.addWidget(info_group)
        
        # Quantité
        quantity_group = QGroupBox("Quantité")
        quantity_layout = QVBoxLayout(quantity_group)
        
        self.quantity_spinbox = QSpinBox()
        self.quantity_spinbox.setRange(1, 99999)
        self.quantity_spinbox.setValue(1)
        
        if operation_type == "OUT" and product:
            self.quantity_spinbox.setMaximum(product.quantity)
        
        quantity_layout.addWidget(QLabel(f"Quantité à {'ajouter' if operation_type == 'IN' else 'retirer'}:"))
        quantity_layout.addWidget(self.quantity_spinbox)
        
        layout.addWidget(quantity_group)
        
        # Prix (seulement pour les entrées)
        if operation_type == "IN":
            price_group = QGroupBox("Prix d'achat")
            price_layout = QVBoxLayout(price_group)
            
            self.price_spinbox = QDoubleSpinBox()
            self.price_spinbox.setRange(0, 99999999.99)
            self.price_spinbox.setDecimals(2)
            self.price_spinbox.setPrefix("FCFA ")
            
            if product and product.purchase_price > 0:
                self.price_spinbox.setValue(product.purchase_price)
            
            price_layout.addWidget(QLabel("Prix d'achat unitaire:"))
            price_layout.addWidget(self.price_spinbox)
            
            layout.addWidget(price_group)
        
        # Notes
        notes_group = QGroupBox("Notes")
        notes_layout = QVBoxLayout(notes_group)
        
        self.notes_text = QTextEdit()
        self.notes_text.setMaximumHeight(100)
        self.notes_text.setPlaceholderText("Raison de l'opération, référence, etc.")
        
        notes_layout.addWidget(self.notes_text)
        layout.addWidget(notes_group)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)


class AdjustStockDialog(QDialog):
    """Boîte de dialogue pour ajuster le stock d'un produit"""
    def __init__(self, parent=None, product=None):
        super().__init__(parent)
        
        self.product = product
        
        self.setWindowTitle("Ajuster le stock")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Informations produit
        info_group = QGroupBox("Informations du produit")
        info_layout = QVBoxLayout(info_group)
        
        if product:
            info_layout.addWidget(QLabel(f"<b>Produit:</b> {product.name}"))
            info_layout.addWidget(QLabel(f"<b>Code:</b> {product.code or 'N/A'}"))
            info_layout.addWidget(QLabel(f"<b>Stock actuel:</b> {product.quantity:,} unités"))
            info_layout.addWidget(QLabel(f"<b>Stock minimum:</b> {product.min_stock:,} unités"))
            info_layout.addWidget(QLabel(f"<b>Stock maximum:</b> {product.max_stock:,} unités"))
        
        layout.addWidget(info_group)
        
        # Nouvelle quantité
        quantity_group = QGroupBox("Nouvelle quantité")
        quantity_layout = QVBoxLayout(quantity_group)
        
        self.quantity_spinbox = QSpinBox()
        self.quantity_spinbox.setRange(0, 99999)
        
        if product:
            self.quantity_spinbox.setValue(product.quantity)
        
        quantity_layout.addWidget(QLabel("Nouvelle quantité en stock:"))
        quantity_layout.addWidget(self.quantity_spinbox)
        
        # Calcul de différence
        self.difference_label = QLabel("Différence: 0")
        quantity_layout.addWidget(self.difference_label)
        
        layout.addWidget(quantity_group)
        
        # Connecter le changement de quantité
        self.quantity_spinbox.valueChanged.connect(self.update_difference)
        
        # Notes
        notes_group = QGroupBox("Notes")
        notes_layout = QVBoxLayout(notes_group)
        
        self.notes_text = QTextEdit()
        self.notes_text.setMaximumHeight(100)
        self.notes_text.setPlaceholderText("Raison de l'ajustement (inventaire, erreur, etc.)")
        
        notes_layout.addWidget(self.notes_text)
        layout.addWidget(notes_group)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def update_difference(self):
        """Mettre à jour l'affichage de la différence"""
        if self.product:
            difference = self.quantity_spinbox.value() - self.product.quantity
            self.difference_label.setText(f"Différence: {'+' if difference > 0 else ''}{difference}")
            
            if difference > 0:
                self.difference_label.setStyleSheet("color: #10b981;")
            elif difference < 0:
                self.difference_label.setStyleSheet("color: #ef4444;")
            else:
                self.difference_label.setStyleSheet("")


class InventoryDialog(QDialog):
    """Boîte de dialogue pour l'inventaire"""
    def __init__(self, parent=None, db_session=None, user=None):
        super().__init__(parent)
        
        self.db_session = db_session
        self.user = user
        
        self.setWindowTitle("Inventaire physique")
        self.setModal(True)
        self.setMinimumSize(500, 400)
        
        layout = QVBoxLayout(self)
        
        # En-tête
        header_label = QLabel("Inventaire physique - Comptage des stocks")
        header_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(header_label)
        
        # Instructions
        instructions = QLabel(
            "Veuillez entrer les quantités comptées pour chaque produit.\n"
            "Les différences seront automatiquement calculées et enregistrées."
        )
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Tableau d'inventaire
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Produit", "Code", "Stock actuel", "Quantité comptée", "Différence", "Notes"
        ])
        
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        
        # Résumé
        self.summary_label = QLabel("Prêt pour l'inventaire")
        layout.addWidget(self.summary_label)
        
        # Boutons
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("Enregistrer l'inventaire")
        save_btn.setIcon(parent.style().standardIcon(QStyle.SP_DialogSaveButton))
        save_btn.clicked.connect(self.save_inventory)
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setIcon(parent.style().standardIcon(QStyle.SP_DialogCancelButton))
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)
        
        layout.addLayout(button_layout)
        
        # Charger les produits
        self.load_products()
    
    def load_products(self):
        """Charger les produits dans le tableau"""
        try:
            products = self.db_session.query(Product).filter(Product.active == True).order_by(Product.name).all()
            
            self.table.setRowCount(len(products))
            
            for row, product in enumerate(products):
                # Produit
                name_item = QTableWidgetItem(product.name)
                self.table.setItem(row, 0, name_item)
                
                # Code
                code_item = QTableWidgetItem(product.code or "")
                self.table.setItem(row, 1, code_item)
                
                # Stock actuel
                current_item = QTableWidgetItem(str(product.quantity))
                current_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, 2, current_item)
                
                # Quantité comptée (spinbox)
                counted_spinbox = QSpinBox()
                counted_spinbox.setRange(0, 99999)
                counted_spinbox.setValue(product.quantity)
                counted_spinbox.valueChanged.connect(lambda value, r=row, p=product: self.update_difference(r, value, p))
                self.table.setCellWidget(row, 3, counted_spinbox)
                
                # Différence
                diff_item = QTableWidgetItem("0")
                diff_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, 4, diff_item)
                
                # Notes (input)
                notes_input = QLineEdit()
                notes_input.setPlaceholderText("Notes (optionnel)")
                self.table.setCellWidget(row, 5, notes_input)
            
            # Mettre à jour le résumé
            self.update_summary()
            
        except Exception as e:
            print(f"Erreur lors du chargement des produits: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")
    
    def update_difference(self, row, counted_value, product):
        """Mettre à jour la différence pour une ligne"""
        try:
            current = product.quantity
            difference = counted_value - current
            
            diff_item = self.table.item(row, 4)
            if diff_item:
                diff_item.setText(str(difference))
                
                if difference > 0:
                    diff_item.setForeground(QColor("#10b981"))
                    diff_item.setToolTip(f"Excédent de {difference} unités")
                elif difference < 0:
                    diff_item.setForeground(QColor("#ef4444"))
                    diff_item.setToolTip(f"Déficit de {abs(difference)} unités")
                else:
                    diff_item.setForeground(QColor("#000000"))
                    diff_item.setToolTip("Aucune différence")
            
            # Mettre à jour le résumé
            self.update_summary()
        except Exception as e:
            print(f"Erreur dans update_difference: {e}")
    
    def update_summary(self):
        """Mettre à jour le résumé de l'inventaire"""
        try:
            total_diff = 0
            items_diff = 0
            
            for row in range(self.table.rowCount()):
                diff_item = self.table.item(row, 4)
                if diff_item:
                    try:
                        diff = int(diff_item.text())
                        if diff != 0:
                            items_diff += 1
                            total_diff += diff
                    except ValueError:
                        continue
            
            self.summary_label.setText(
                f"Produits à ajuster: {items_diff} | "
                f"Différence totale: {'+' if total_diff > 0 else ''}{total_diff:,} unités"
            )
        except Exception as e:
            print(f"Erreur dans update_summary: {e}")
    
    def save_inventory(self):
        """Sauvegarder les résultats de l'inventaire"""
        try:
            user_id = self.user.get('id') if isinstance(self.user, dict) else self.user.id
            
            for row in range(self.table.rowCount()):
                product_name_item = self.table.item(row, 0)
                product_code_item = self.table.item(row, 1)
                
                if not product_name_item:
                    continue
                    
                product_name = product_name_item.text()
                product_code = product_code_item.text() if product_code_item else ""
                
                # Trouver le produit
                product = None
                if product_code:
                    product = self.db_session.query(Product).filter(Product.code == product_code).first()
                
                if not product and product_name:
                    product = self.db_session.query(Product).filter(Product.name == product_name).first()
                
                if product:
                    counted_spinbox = self.table.cellWidget(row, 3)
                    diff_item = self.table.item(row, 4)
                    notes_input = self.table.cellWidget(row, 5)
                    
                    if counted_spinbox and diff_item:
                        new_quantity = counted_spinbox.value()
                        difference = int(diff_item.text())
                        notes = notes_input.text() if notes_input else ""
                        
                        if difference != 0:
                            # Créer un mouvement d'ajustement
                            movement = InventoryMovement(
                                product_id=product.id,
                                movement_type="ADJUST",
                                quantity=abs(difference),
                                unit_price=product.purchase_price,
                                total_value=abs(difference) * product.purchase_price,
                                reason="Inventaire physique",
                                notes=f"{notes}\nInventaire: {product.quantity} → {new_quantity} (différence: {'+' if difference > 0 else ''}{difference})",
                                user_id=user_id
                            )
                            
                            self.db_session.add(movement)
                            
                            # Mettre à jour la quantité
                            product.quantity = new_quantity
            
            self.db_session.commit()
            QMessageBox.information(self, "Succès", "Inventaire enregistré avec succès!")
            self.accept()
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'enregistrement: {str(e)}")


class NewMovementDialog(QDialog):
    """Boîte de dialogue pour nouveau mouvement"""
    def __init__(self, parent=None, db_session=None, user=None):
        super().__init__(parent)
        
        self.db_session = db_session
        self.user = user
        
        self.setWindowTitle("Nouveau mouvement de stock")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        # Formulaire
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        
        # Produit
        self.product_combo = QComboBox()
        self.product_combo.currentIndexChanged.connect(self.on_product_changed)
        
        # Charger les produits
        products = self.db_session.query(Product).filter(Product.active == True).order_by(Product.name).all()
        for product in products:
            stock_info = f" ({product.quantity:,} en stock)" if product.quantity > 0 else " (rupture)"
            self.product_combo.addItem(f"{product.code or 'N/A'} - {product.name}{stock_info}", product.id)
        
        form.addRow("Produit*:", self.product_combo)
        
        # Type de mouvement
        self.type_combo = QComboBox()
        self.type_combo.addItem("IN - Entrée de stock", "IN")
        self.type_combo.addItem("OUT - Sortie de stock", "OUT")
        self.type_combo.addItem("ADJUST - Ajustement", "ADJUST")
        self.type_combo.addItem("LOSS - Perte", "LOSS")
        self.type_combo.addItem("RETURN - Retour", "RETURN")
        self.type_combo.currentIndexChanged.connect(self.on_type_changed)
        
        form.addRow("Type*:", self.type_combo)
        
        # Quantité
        self.quantity_spinbox = QSpinBox()
        self.quantity_spinbox.setRange(1, 99999)
        self.quantity_spinbox.setValue(1)
        
        form.addRow("Quantité*:", self.quantity_spinbox)
        
        # Prix unitaire
        self.price_label = QLabel("Prix unitaire:")
        self.price_spinbox = QDoubleSpinBox()
        self.price_spinbox.setRange(0, 99999999.99)
        self.price_spinbox.setDecimals(2)
        self.price_spinbox.setPrefix("FCFA ")
        
        form.addRow(self.price_label, self.price_spinbox)
        
        # Date
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        
        form.addRow("Date:", self.date_edit)
        
        # Raison
        self.reason_combo = QComboBox()
        self.reason_combo.addItems([
            "Achat fournisseur",
            "Vente client",
            "Transfert",
            "Inventaire",
            "Dons/Cadeaux",
            "Échantillons",
            "Détérioration",
            "Autre"
        ])
        self.reason_combo.setEditable(True)
        
        form.addRow("Raison:", self.reason_combo)
        
        # Référence
        self.reference_input = QLineEdit()
        self.reference_input.setPlaceholderText("Numéro de facture, commande, etc.")
        
        form.addRow("Référence:", self.reference_input)
        
        # Notes
        self.notes_text = QTextEdit()
        self.notes_text.setMaximumHeight(80)
        self.notes_text.setPlaceholderText("Notes complémentaires...")
        
        form.addRow("Notes:", self.notes_text)
        
        layout.addLayout(form)
        
        # Informations produit
        self.product_info_label = QLabel("")
        self.product_info_label.setWordWrap(True)
        layout.addWidget(self.product_info_label)
        
        # Mettre à jour les informations initiales
        self.on_product_changed()
        self.on_type_changed()
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def on_product_changed(self):
        """Mettre à jour les informations du produit"""
        product_id = self.product_combo.currentData()
        if product_id:
            product = self.db_session.query(Product).get(product_id)
            if product:
                info = f"<b>{product.name}</b><br>"
                info += f"Stock actuel: {product.quantity:,} unités<br>"
                info += f"Prix d'achat: {product.purchase_price:,.2f} FCFA<br>"
                info += f"Prix de vente: {product.sale_price:,.2f} FCFA"
                
                if product.quantity <= product.min_stock:
                    info += f"<br><span style='color: #ef4444;'>⚠️ Stock faible (minimum: {product.min_stock})</span>"
                
                self.product_info_label.setText(info)
                
                # Définir le prix par défaut
                if self.type_combo.currentData() in ["IN", "ADJUST"]:
                    self.price_spinbox.setValue(product.purchase_price)
    
    def on_type_changed(self):
        """Mettre à jour l'interface selon le type"""
        movement_type = self.type_combo.currentData()
        
        # Afficher/masquer le prix
        show_price = movement_type in ["IN", "ADJUST"]
        self.price_label.setVisible(show_price)
        self.price_spinbox.setVisible(show_price)
        
        # Mettre à jour le produit sélectionné
        self.on_product_changed()
    
    def validate_and_save(self):
        """Valider et sauvegarder le mouvement"""
        if self.product_combo.currentIndex() == -1:
            QMessageBox.warning(self, "Validation", "Veuillez sélectionner un produit!")
            return
        
        if self.quantity_spinbox.value() <= 0:
            QMessageBox.warning(self, "Validation", "La quantité doit être supérieure à 0!")
            return
        
        movement_type = self.type_combo.currentData()
        product_id = self.product_combo.currentData()
        product = self.db_session.query(Product).get(product_id)
        
        # Vérifier le stock pour les sorties
        if movement_type == "OUT" and product:
            if self.quantity_spinbox.value() > product.quantity:
                QMessageBox.warning(self, "Validation", 
                                  f"Quantité insuffisante!\n"
                                  f"Stock disponible: {product.quantity:,} unités")
                return
        
        try:
            user_id = self.user.get('id') if isinstance(self.user, dict) else self.user.id
            
            movement = InventoryMovement(
                product_id=product_id,
                movement_type=movement_type,
                quantity=self.quantity_spinbox.value(),
                unit_price=self.price_spinbox.value() if movement_type in ["IN", "ADJUST"] else None,
                total_value=self.quantity_spinbox.value() * self.price_spinbox.value() if movement_type in ["IN", "ADJUST"] else None,
                reason=self.reason_combo.currentText(),
                reference=self.reference_input.text().strip() or None,
                notes=self.notes_text.toPlainText().strip() or None,
                user_id=user_id,
                date=self.date_edit.date().toPython()
            )
            
            self.db_session.add(movement)
            
            # Mettre à jour la quantité du produit
            if movement_type == "IN":
                product.quantity += self.quantity_spinbox.value()
            elif movement_type == "OUT":
                product.quantity -= self.quantity_spinbox.value()
            elif movement_type == "ADJUST":
                product.quantity += self.quantity_spinbox.value()
            
            self.db_session.commit()
            QMessageBox.information(self, "Succès", "Mouvement enregistré!")
            self.accept()
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")


class ExpenseDialog(QDialog):
    """Boîte de dialogue pour nouvelle dépense"""
    def __init__(self, parent=None, db_session=None, user=None):
        super().__init__(parent)
        
        self.db_session = db_session
        self.user = user
        
        self.setWindowTitle("Nouvelle dépense")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        # Formulaire
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        
        # Catégorie
        self.category_combo = QComboBox()
        
        # Charger les catégories
        categories = self.db_session.query(ExpenseCategory).filter(ExpenseCategory.active == True).order_by(ExpenseCategory.name).all()
        for category in categories:
            self.category_combo.addItem(category.name, category.id)
        
        self.category_combo.setEditable(True)
        form.addRow("Catégorie*:", self.category_combo)
        
        # Montant
        self.amount_spinbox = QDoubleSpinBox()
        self.amount_spinbox.setRange(0, 99999999.99)
        self.amount_spinbox.setDecimals(2)
        self.amount_spinbox.setPrefix("FCFA ")
        
        form.addRow("Montant*:", self.amount_spinbox)
        
        # Description
        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("Description de la dépense")
        
        form.addRow("Description*:", self.description_input)
        
        # Date
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        
        form.addRow("Date:", self.date_edit)
        
        # Moyen de paiement
        self.payment_combo = QComboBox()
        self.payment_combo.addItems(["ESPÈCES", "VIREMENT BANCAIRE", "MOBILE MONEY", "CARTE BANCAIRE", "CHÈQUE", "AUTRE"])
        
        form.addRow("Moyen de paiement:", self.payment_combo)
        
        # Référence
        self.reference_input = QLineEdit()
        self.reference_input.setPlaceholderText("Numéro de facture, reçu, etc.")
        
        form.addRow("Référence:", self.reference_input)
        
        # Fournisseur
        self.supplier_combo = QComboBox()
        self.supplier_combo.addItem("-- Non spécifié --", None)
        
        # Charger les fournisseurs
        suppliers = self.db_session.query(Supplier).filter(Supplier.active == True).order_by(Supplier.name).all()
        for supplier in suppliers:
            self.supplier_combo.addItem(supplier.name, supplier.id)
        
        form.addRow("Fournisseur:", self.supplier_combo)
        
        # Notes
        self.notes_text = QTextEdit()
        self.notes_text.setMaximumHeight(80)
        self.notes_text.setPlaceholderText("Notes complémentaires...")
        
        form.addRow("Notes:", self.notes_text)
        
        layout.addLayout(form)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def validate_and_save(self):
        """Valider et sauvegarder la dépense"""
        if self.category_combo.currentText().strip() == "":
            QMessageBox.warning(self, "Validation", "Veuillez sélectionner ou entrer une catégorie!")
            return
        
        if self.amount_spinbox.value() <= 0:
            QMessageBox.warning(self, "Validation", "Le montant doit être supérieur à 0!")
            return
        
        if self.description_input.text().strip() == "":
            QMessageBox.warning(self, "Validation", "La description est obligatoire!")
            return
        
        try:
            user_id = self.user.get('id') if isinstance(self.user, dict) else self.user.id
            
            # Vérifier la catégorie
            category_name = self.category_combo.currentText().strip()
            category = self.db_session.query(ExpenseCategory).filter(ExpenseCategory.name == category_name).first()
            
            if not category:
                category = ExpenseCategory(name=category_name)
                self.db_session.add(category)
                self.db_session.flush()
            
            # Créer la dépense
            expense = Expense(
                category_id=category.id,
                amount=self.amount_spinbox.value(),
                description=self.description_input.text().strip(),
                payment_method=self.payment_combo.currentText(),
                reference=self.reference_input.text().strip() or None,
                supplier_id=self.supplier_combo.currentData(),
                user_id=user_id,
                date=self.date_edit.date().toPython()
            )
            
            self.db_session.add(expense)
            self.db_session.commit()
            
            QMessageBox.information(self, "Succès", "Dépense enregistrée avec succès!")
            self.accept()
            
        except Exception as e:
            self.db_session.rollback()
            QMessageBox.critical(self, "Erreur", f"Erreur: {str(e)}")


class ProfitCalculationDialog(QDialog):
    """Boîte de dialogue pour calculer les bénéfices"""
    def __init__(self, parent=None, products=None, db_session=None):
        super().__init__(parent)
        
        self.products = products
        self.db_session = db_session
        self.parent_view = parent
        
        self.setWindowTitle("Calcul détaillé des bénéfices")
        self.setModal(True)
        self.setMinimumSize(500, 400)
        self.setModal(True)
        self.setMinimumSize(500, 400)
        
        layout = QVBoxLayout(self)
        
        # Calcul des bénéfices
        total_investment = sum(p.stock_value for p in products) if products else 0
        total_potential = sum(p.quantity * p.sale_price for p in products) if products else 0
        total_profit = total_potential - total_investment
        
        # Dépenses totales
        all_expenses = db_session.query(Expense).all() if db_session else []
        total_expenses = sum(e.amount for e in all_expenses)
        net_profit = total_profit - total_expenses
        
        # Affichage des résultats
        currency = parent.currency if hasattr(parent, 'currency') else "FCFA"
        
        results_html = f"""
        <h3 style='text-align: center;'>Calcul Détaillé des Bénéfices</h3>
        
        <table style='border-collapse: collapse; width: 100%; margin: 20px 0;'>
            <tr style='background-color: #f8fafc;'>
                <td style='padding: 12px; font-weight: bold; border: 1px solid #e2e8f0;'>Investissement total en stock:</td>
                <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; font-weight: bold;'>{total_investment:,.2f} {currency}</td>
            </tr>
            <tr>
                <td style='padding: 12px; border: 1px solid #e2e8f0;'>Valeur potentielle de vente:</td>
                <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0;'>{total_potential:,.2f} {currency}</td>
            </tr>
            <tr style='background-color: #f0fdf4;'>
                <td style='padding: 12px; font-weight: bold; border: 1px solid #e2e8f0; color: #15803d;'>Profit brut potentiel:</td>
                <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; font-weight: bold; color: #15803d;'>{total_profit:,.2f} {currency}</td>
            </tr>
            <tr>
                <td style='padding: 12px; border: 1px solid #e2e8f0;'>Dépenses totales enregistrées:</td>
                <td style='padding: 12px; text-align: right; border: 1px solid #e2e8f0; color: #dc2626;'>{total_expenses:,.2f} {currency}</td>
            </tr>
            <tr style='background-color: #dbeafe; border-top: 3px solid #3b82f6;'>
                <td style='padding: 15px; font-weight: bold; font-size: 16px; border: 1px solid #e2e8f0;'>BÉNÉFICE NET ESTIMÉ:</td>
                <td style='padding: 15px; text-align: right; font-weight: bold; font-size: 16px; color: #1d4ed8; border: 1px solid #e2e8f0;'>
                    {net_profit:,.2f} {currency}
                </td>
            </tr>
        </table>

        <div style='margin: 20px 0; padding: 15px; background-color: #f8fafc; border-radius: 8px;'>
            <p><b>Taux de marge:</b> {(total_profit/total_investment*100) if total_investment > 0 else 0:.1f}%</p>
            <p><b>Retour sur investissement (ROI):</b> {(net_profit/total_investment*100) if total_investment > 0 else 0:.1f}%</p>
            <p><b>Marge bénéficiaire nette:</b> {(net_profit/total_potential*100) if total_potential > 0 else 0:.1f}%</p>
        </div>
        
        <h4 style='margin-top: 20px;'>Analyse par marge:</h4>
        <ul>
            <li>Produits à haute marge (> 30%): {len([p for p in products if p.profit_margin > 30]) if products else 0:,}</li>
            <li>Produits à marge moyenne (10-30%): {len([p for p in products if 10 <= p.profit_margin <= 30]) if products else 0:,}</li>
            <li>Produits à faible marge (< 10%): {len([p for p in products if p.profit_margin < 10]) if products else 0:,}</li>
        </ul>
        """
        
        results_label = QLabel(results_html)
        results_label.setWordWrap(True)
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(results_label)
        scroll_area.setWidgetResizable(True)
        
        layout.addWidget(scroll_area)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.reject)
        
        # Bouton d'export
        export_btn = QPushButton("Exporter le rapport en Excel")
        export_btn.setIcon(parent.style().standardIcon(QStyle.SP_DialogSaveButton))
        export_btn.clicked.connect(lambda: self.export_report_excel(total_investment, total_potential, total_profit, total_expenses, net_profit))
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(export_btn)
        button_layout.addStretch()
        button_layout.addWidget(buttons)
        
        layout.addLayout(button_layout)
    
    def export_report_excel(self, total_investment, total_potential, total_profit, total_expenses, net_profit):
        """Exporter le rapport de bénéfices en Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
            from PySide6.QtWidgets import QFileDialog
            from datetime import datetime
            
            default_name = f"calcul_benefices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter le calcul de bénéfices", default_name,
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*.*)"
            )
            
            if not filename:
                return
            
            currency = self.parent_view.currency if hasattr(self.parent_view, 'currency') else "FCFA"
            
            # Créer le contenu Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Données principales
                main_data = {
                    'ÉLÉMENT': [
                        'Investissement total en stock',
                        'Valeur potentielle de vente',
                        'Profit brut potentiel',
                        'Dépenses totales enregistrées',
                        'BÉNÉFICE NET ESTIMÉ'
                    ],
                    f'MONTANT ({currency})': [
                        total_investment,
                        total_potential,
                        total_profit,
                        total_expenses,
                        net_profit
                    ],
                    'POURCENTAGE': [
                        '100.0%',
                        f'{(total_potential/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(total_profit/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(total_expenses/total_investment*100) if total_investment > 0 else 0:.1f}%',
                        f'{(net_profit/total_investment*100) if total_investment > 0 else 0:.1f}%'
                    ]
                }
                
                df_main = pd.DataFrame(main_data)
                df_main.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Analyse par marge
                if self.products:
                    high_margin = [p for p in self.products if p.profit_margin > 30]
                    medium_margin = [p for p in self.products if 10 <= p.profit_margin <= 30]
                    low_margin = [p for p in self.products if p.profit_margin < 10]
                    
                    margin_data = {
                        'Catégorie': [
                            'Haute marge (> 30%)',
                            'Marge moyenne (10-30%)',
                            'Faible marge (< 10%)',
                            'TOTAL'
                        ],
                        'Nombre de produits': [
                            len(high_margin),
                            len(medium_margin),
                            len(low_margin),
                            len(self.products)
                        ],
                        f'Valeur du stock ({currency})': [
                            sum(p.stock_value for p in high_margin),
                            sum(p.stock_value for p in medium_margin),
                            sum(p.stock_value for p in low_margin),
                            total_investment
                        ],
                        'Pourcentage': [
                            f'{(sum(p.stock_value for p in high_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                            f'{(sum(p.stock_value for p in medium_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                            f'{(sum(p.stock_value for p in low_margin)/total_investment*100) if total_investment > 0 else 0:.1f}%',
                            '100.0%'
                        ]
                    }
                    
                    df_margin = pd.DataFrame(margin_data)
                    df_margin.to_excel(writer, sheet_name='Analyse par marge', index=False)
                    
                    # Détail des produits
                    products_data = []
                    for product in self.products:
                        supplier_name = product.supplier.name if product.supplier else ""
                        margin_category = "Haute" if product.profit_margin > 30 else "Moyenne" if product.profit_margin >= 10 else "Faible"
                        
                        products_data.append({
                            'Code': product.code or "",
                            'Nom': product.name,
                            'Catégorie': product.category,
                            'Quantité': product.quantity,
                            'Prix Achat': product.purchase_price,
                            'Prix Vente': product.sale_price,
                            f'Valeur Stock ({currency})': product.stock_value,
                            'Marge %': f"{product.profit_margin:.1f}%",
                            'Catégorie Marge': margin_category,
                            'Fournisseur': supplier_name,
                            'Actif': "Oui" if product.active else "Non"
                        })
                    
                    df_products = pd.DataFrame(products_data)
                    df_products.to_excel(writer, sheet_name='Détail Produits', index=False)
                
                # Formater automatiquement les colonnes
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formater les colonnes monétaires
                    if sheet_name == 'Résumé':
                        for cell in worksheet['B'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Analyse par marge':
                        for cell in worksheet['C'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    elif sheet_name == 'Détail Produits':
                        for cell in worksheet['F'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['G'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        for cell in worksheet['H'][1:]:
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            QMessageBox.information(self, "Export réussi", f"Rapport exporté vers:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export: {str(e)}")
            import traceback
            traceback.print_exc()